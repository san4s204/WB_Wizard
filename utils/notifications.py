from collections import defaultdict
from aiogram import Bot

from parse_wb import parse_wildberries
from db.database import SessionLocal
from sqlalchemy import func, desc
from db.models import Order, ReportDetails, Stock, User, Product, UserWarehouse, Token
from db.models import Sale
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import io
import datetime
from datetime import timedelta


def calc_price_with_spp(finished_price: float, spp: float) -> float:
    """
    finished_price: базовая цена товара
    spp: скидка (проценты), например 20.0 => 20%
    Возвращаем итоговую цену с учётом скидки
    """
    discount_amount = finished_price * (spp / 100.0)
    final_price = finished_price - discount_amount
    return final_price

def get_average_daily_orders(nm_id: int, days=90) -> float:
    """
    Возвращает среднее кол-во заказов (Order) в сутки за последние X дней 
    для товара nm_id.
    """
    session = SessionLocal()
    date_from = datetime.date.today() - datetime.timedelta(days=days)
    # сколько заказов (is_cancel=False) за этот период
    total_orders = session.query(func.count(Order.id)) \
        .filter(Order.nm_id == nm_id) \
        .filter(Order.date >= date_from) \
        .filter(Order.is_cancel == False) \
        .scalar() or 0
    session.close()

    avg_per_day = total_orders / days if days > 0 else 0
    return avg_per_day


def get_average_daily_sales(nm_id: int, days=90) -> float:
    """
    Возвращает среднее кол-во заказов (Order) в сутки за последние X дней 
    для товара nm_id.
    """
    session = SessionLocal()
    date_from = datetime.date.today() - datetime.timedelta(days=days)
    # сколько заказов (is_cancel=False) за этот период
    total_sales = session.query(func.count(Sale.id)) \
        .filter(Sale.nm_id == nm_id) \
        .filter(Sale.date >= date_from) \
        .scalar() or 0
    session.close()

    avg_per_day = total_sales / days if days > 0 else 0
    return avg_per_day

def count_today_orders_by_nmId(nm_id: int) -> int:
    session = SessionLocal()
    today = datetime.date.today()
    result = session.query(func.count(Order.id))\
        .filter(Order.nm_id == nm_id)\
        .filter(func.date(Order.date) == today)\
        .scalar()
    session.close()
    return result

def get_latest_delivery_cost(nm_id: int, office_name: str) -> float:
    """
    Возвращает стоимость логистики (delivery_rub) из последней записи report_details
    для заданного nm_id и office_name.
    """
    session = SessionLocal()
    try:
        latest_record = (
            session.query(ReportDetails.delivery_rub)
            .filter(ReportDetails.nm_id == nm_id)
            .filter(ReportDetails.office_name == office_name)
            .order_by(desc(ReportDetails.order_dt))
            .first()
        )
        return latest_record.delivery_rub if latest_record else 0.0
    finally:
        session.close()

def get_latest_commision(nm_id: int) -> int:
    """
    Возвращает комиссию (commission) из последней записи report_details для заданного nm_id.
    """
    session = SessionLocal()
    try:
        latest_record = (
            session.query(ReportDetails.commission_percent)
            .filter(ReportDetails.nm_id == nm_id)
            .filter(ReportDetails.quantity > 0)
            .order_by(desc(ReportDetails.order_dt))
            .first()
        )
        return latest_record.commission_percent if latest_record else 0
    finally:
        session.close()

def get_total_stock(nm_id: int) -> int:
    """
    Возвращает суммарный остаток товара (quantity) по всем складам для заданного nm_id.
    """
    session = SessionLocal()
    try:
        total_quantity = (
            session.query(func.sum(Stock.quantity))
            .filter(Stock.nm_id == nm_id)
            .scalar()
        )
        return total_quantity if total_quantity else 0
    finally:
        session.close()

def get_orders_last_3_months(nm_id: int) -> int:
    """
    Возвращает количество заказов за последние 3 месяца для nm_id.
    """
    session = SessionLocal()
    try:
        three_months_ago = datetime.date.today() - datetime.timedelta(days=90)
        count =( 
            session.query(func.count(Order.id))
            .filter(Order.nm_id == nm_id)
            .filter(Order.date >= three_months_ago)
            .scalar()
        )
        return count if count else 0
    finally:
        session.close()

def get_sales_last_3_months(nm_id: int) -> int:
    """
    Возвращает количество заказов за последние 3 месяца для nm_id.
    """
    session = SessionLocal()
    try:
        three_months_ago = datetime.date.today() - datetime.timedelta(days=90)
        count =( 
            session.query(func.count(Sale.id))
            .filter(Sale.nm_id == nm_id)
            .filter(Sale.date >= three_months_ago)
            .scalar()
        )
        return count if count else 0
    finally:
        session.close()


def count_today_sales_by_nmId(nm_id: int) -> int:
    """
    Считает количество выкупов (sale) по nm_id за сегодня.
    """
    session = SessionLocal()
    today = datetime.date.today()
    result = session.query(func.count(Sale.id)) \
        .filter(Sale.nm_id == nm_id) \
        .filter(func.date(Sale.date) == today) \
        .scalar()
    session.close()
    return result

async def notify_new_orders(bot: Bot, orders_data: list[dict]):

    """
    Отправляет уведомление о заказах - теперь не CSV, а сразу формируем текстовое сообщение:
    Дата: 
    Товар: 
    Артикул:
    Рейтинг:
    Отзывы:
    Отгрузка:
    Доставка:
    Сегодня:
    Цена с СПП:
    """

    if not orders_data:
        return

    # Группируем заказы по token_id
    grouped_orders = defaultdict(list)
    for order in orders_data:
        tid = order.get("token_id")
        grouped_orders[tid].append(order)

    session = SessionLocal()

    # Для каждого token_id достаём пользователей, рассылаем
    for token_id, orders_list in grouped_orders.items():

        # Ищем пользователей, у которых user.token_id == token_id
        # и notify_orders=True
        users = session.query(User).filter(
            User.token_id == token_id,
            User.notify_orders == True
        ).all()

        if not users:
            continue  # Никто не подписан на этот токен или нет таких пользователей

        for order in orders_list:
            # Формируем сообщение (пример, как было раньше)
            nm_id = order.get("nm_id")
            url = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"
            item_name = order.get("itemName", "N/A")
            base_price = float(order.get("price_with_disc", 0.0))
            spp_value = float(order.get("spp", 0.0))
            final_price = calc_price_with_spp(base_price, spp_value)
            rating = order.get("rating", "N/A")
            reviews = order.get("reviews", "N/A")
            picture_url = order.get("image_url", None)

            date_str = order.get("date", "N/A")
            date_str = date_str.replace("T", " ")

            warehouse_name = order.get("warehouseName", "N/A")
            region_name = order.get("regionName", "N/A")
            delivery_cost = get_latest_delivery_cost(nm_id, warehouse_name)
            today_count = count_today_orders_by_nmId(nm_id)
            orders_last_3_months = get_orders_last_3_months(nm_id)
            total_stocks = get_total_stock(nm_id)
            avg_daily_usage = get_average_daily_orders(nm_id, days=90)  # например, 30 дней
            days_coverage = total_stocks / avg_daily_usage if avg_daily_usage > 0 else 0
            delivery_rub = get_latest_delivery_cost(nm_id, warehouse_name)

            caption_text = (
                f"🆕🛍 <b>Новый заказ!</b>\n"
                f"📅 <b>Дата:</b> {date_str}\n"
                f"📦 <b>Товар:</b> {item_name}\n"
                f"🔖 <b>Артикул:</b> <a href='{url}'>{nm_id}</a>\n"
                f"⭐ <b>Рейтинг:</b> {rating}\n"
                f"💬 <b>Отзывы:</b> {reviews}\n"
                f"🚚 <b>Отгрузка:</b> {warehouse_name}\n"
                f"💰 <b>Логистика:</b> {delivery_rub:.2f}\n"
                f"🏙 <b>Доставка:</b> {region_name}\n"
                f"💲 <b>Сумма:</b> {base_price:.2f}  |  🔽 <b>Цена с СПП:</b> {final_price:.2f}\n"
                f"📆 <b>Сегодня:</b> {today_count}\n"
                f"📊 <b>Заказов за 3 месяца:</b> {orders_last_3_months}\n"
                f"\n"
                f"📦 <b>Остаток:</b> {total_stocks} шт. ⏳ <b>Хватит примерно на:</b> {days_coverage:.0f} дн."
            )

            # Рассылаем всем пользователям, у которых token_id == token_id
            for user in users:
                chat_id = user.telegram_id
                try:
                    if picture_url:
                        # Пытаемся отправить фото
                        try:
                            await bot.send_photo(
                                chat_id=chat_id,
                                photo=picture_url,
                                caption=caption_text,
                                parse_mode="HTML"
                            )
                        except Exception:
                            fallback_text = f"{picture_url}\n{caption_text}"
                            await bot.send_message(chat_id=chat_id, text=fallback_text)
                    else:
                        await bot.send_message(chat_id=chat_id, text=caption_text, parse_mode="HTML")
                except Exception as e:
                    print(f"Ошибка при отправке пользователю {chat_id}: {e}")

    session.close()
    print("Уведомления о новых заказах отправлены!")

async def notify_free_incomes(bot: Bot, incomes_data: list[dict]):
    """
    Отправляет уведомления о бесплатных поставках (totalPrice=0) в формате:
      Обнаружена бесплатная поставка!
      Дата: <дата первой позиции>
      Склад: <название склада>
      Товары:
        - {subject_name_1} x {quantity_1}
        - {subject_name_2} x {quantity_2}
      Всего товаров: {итоговое количество}
    """

    if not incomes_data:
        return



    # 1) Сгруппируем поступления по token_id
    token_groups = defaultdict(list)
    for inc in incomes_data:
        tid = inc.get("token_id")
        token_groups[tid].append(inc)

    session = SessionLocal()

    for token_id, inc_list in token_groups.items():
        # 2) Ищем пользователей, у которых token_id == token_id
        #    и, например, notify_orders == True (или другой флаг)
        users = session.query(User).filter(
            User.token_id == token_id,
            User.notify_incomes == True  # <-- или свой флаг notify_incomes
        ).all()
        if not users:
            continue  # Нет подписанных пользователей

        # 3) Дополнительная группировка внутри inc_list — по incomeId
        incomeid_groups = defaultdict(list)
        for item in inc_list:
            income_id = item.get("incomeId")
            incomeid_groups[income_id].append(item)

        # 4) Для каждого incomeId собираем «суммарную» поставку
        for income_id, incomes_same_id in incomeid_groups.items():
            # Проверим, «бесплатная» ли (totalPrice=0 у всех строк)
            # Или, если бывает частично платно, частично нет — берём сумму
            total_price_sum = sum(i.get("totalPrice", 0) for i in incomes_same_id)
            if total_price_sum != 0:
                continue  # Это не бесплатная поставка

            # Определяем дату — берём самую раннюю
            # incomes_same_id могут содержать date="2025-03-09T00:00:00"
            # Парсим и берём min
            parsed_dates = []
            for d in incomes_same_id:
                dt_str = d.get("date")
                if dt_str:
                    try:
                        # '2025-03-09T00:00:00'
                        dt = datetime.datetime.fromisoformat(dt_str.replace("Z", ""))
                        parsed_dates.append(dt)
                    except Exception:
                        pass
            if parsed_dates:
                earliest_date = min(parsed_dates)
            else:
                earliest_date = None

            # Склад, берём из первой строки
            warehouse_name = incomes_same_id[0].get("warehouseName", "N/A")

            # Собираем товары
            # subject_names_by_nm = { ... }
            items_info = []
            total_qty = 0

            for line in incomes_same_id:
                nm_id = line.get("nmId")
                qty = line.get("quantity", 0)
                total_qty += qty

                # Получаем subject_name из таблицы Product (если есть)
                product = session.query(Product).filter_by(nm_id=nm_id).first()
                subject_name = product.subject_name if product else f"Товар {nm_id}"

                items_info.append((subject_name, qty))

            # 5) Формируем текст уведомления
            date_str = earliest_date.strftime("%Y-%m-%d") if earliest_date else "N/A"
            text_lines = [
                "🚚 <b>Обнаружена бесплатная поставка!</b>",
                f"📅 <b>Дата:</b> {date_str}",
                f"🏬 <b>Склад:</b> {warehouse_name}",
                "🔹 <b>Товары:</b>"
            ]
            for subj_name, q in items_info:
                text_lines.append(f"   • {subj_name} × {q}")

            text_lines.append(f"\n<b>Всего товаров:</b> {total_qty}")
            msg_text = "\n".join(text_lines)

            # 6) Рассылаем всем пользователям
            for user in users:
                chat_id = user.telegram_id
                try:
                    await bot.send_message(chat_id=chat_id, text=msg_text, parse_mode="HTML")
                except Exception as e:
                    print(f"Ошибка при отправке пользователю {chat_id}: {e}")

    session.close()
    print("Уведомления о бесплатных поставках отправлены.")

async def notify_free_acceptance(bot: Bot, new_coeffs: list[dict]):
    """
    Рассылает уведомления о бесплатной приёмке (coefficient=0),
    ТОЛЬКО для подписанных на склад пользователей.
    """
    if not new_coeffs:
        return

    grouped_by_token = defaultdict(list)
    for item in new_coeffs:
        grouped_by_token[item["token_id"]].append(item)

    session = SessionLocal()

    for token_id, coeff_list in grouped_by_token.items():
        # 1) Пользователи с token_id
        users = session.query(User).filter_by(token_id=token_id).all()
        if not users:
            continue

        # 2) Для каждого нового/обновлённого коэффициента
        for c in coeff_list:
            # Сразу проверяем coefficient == 0
            if c.get("coefficient") != 0:
                continue

            warehouse_id = c.get("warehouse_id")
            if not warehouse_id:
                # Если нет warehouse_id, пропускаем (не отправляем никому)
                continue

            # Парсим дату для красивого вывода
            date_str = c.get("date")
            if date_str:
                try:
                    dt = datetime.datetime.fromisoformat(date_str.replace("Z", ""))
                    date_formatted = dt.strftime("%Y-%m-%d %H:%M")
                except ValueError:
                    date_formatted = date_str
            else:
                date_formatted = "N/A"

            warehouse_name = c.get("warehouse_name", "N/A")
            box_type_name = c.get("box_type_name", "N/A")

            text_lines = [
                "🆓🔔 <b>БЕСПЛАТНАЯ ПРИЁМКА!</b>",
                f"📅 <b>Дата:</b> {date_formatted}",
                f"🏬 <b>Склад:</b> {warehouse_name}",
                f"📦 <b>Тип коробки:</b> {box_type_name}",
                "Коэффициент: Бесплатная",
                f"(данные актуальны на {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')})"
            ]
            msg_text = "\n".join(text_lines)

            # 3) Проверяем, кто подписан на warehouse_id
            user_ids_subscribed = session.query(UserWarehouse.user_id).filter_by(warehouse_id=warehouse_id).all()
            user_ids_subscribed = {row[0] for row in user_ids_subscribed}

            # Среди всех users с token_id оставим только тех, чьи id в user_ids_subscribed
            target_users = [u for u in users if u.id in user_ids_subscribed]
            if not target_users:
                # Никто не подписан, пропускаем
                continue

            # 4) Отправляем
            for user_obj in target_users:
                chat_id = user_obj.telegram_id
                try:
                    await bot.send_message(chat_id=chat_id, text=msg_text, parse_mode="HTML")
                except Exception as exc:
                    print(f"Ошибка при отправке пользователю {chat_id}: {exc}")

    session.close()
    print("Уведомления о бесплатной приёмке отправлены.")

async def notify_new_sales(bot: Bot, sales_data: list[dict]):
    """
    Отправляет уведомление о новых/обновлённых выкупов.
    Аналог notify_new_orders, но для данных from check_new_sales().
    """
    if not sales_data:
        return

    from collections import defaultdict
    grouped_by_token = defaultdict(list)
    for sale in sales_data:
        t_id = sale.get("token_id")
        grouped_by_token[t_id].append(sale)

    session = SessionLocal()

    for token_id, sales_list in grouped_by_token.items():
        # Ищем пользователей, у кого user.token_id == token_id
        users = session.query(User).filter_by(
            token_id=token_id,
            notify_sales=True
        ).all()

        # Если нет пользователей с этим token_id, пропускаем
        if not users:
            continue

        for sale in sales_list:
            nm_id = sale.get("nm_id")
            date_str = (sale.get("date") or "N/A").replace("T", " ")
            item_name = sale.get("itemName", "N/A")
            warehouse_name = sale.get("warehouseName", "N/A")
            region_name = sale.get("regionName", "N/A")

            base_price = float(sale.get("price_with_disc", 0.0))
            spp_value = float(sale.get("spp", 0.0))
            final_price = calc_price_with_spp(base_price, spp_value)
            commision = get_latest_commision(nm_id)

            today_count = count_today_sales_by_nmId(nm_id)
            sales_last_3_months = get_sales_last_3_months(nm_id)
            total_stocks = get_total_stock(nm_id)
            avg_daily_usage = get_average_daily_sales(nm_id, days=90)  # например, 90 дней
            days_coverage = total_stocks / avg_daily_usage if avg_daily_usage > 0 else 0
            delivery_rub = get_latest_delivery_cost(nm_id, warehouse_name)

            rating = sale.get("rating", "N/A")
            reviews = sale.get("reviews", "N/A")
            image_url = sale.get("image_url", None)

            nm_id_link = f"<a href='https://www.wildberries.ru/catalog/{nm_id}/detail.aspx'>{nm_id}</a>"

            caption_text = (
                f"🆕🔔 <b>Новый выкуп!</b>\n"
                f"📅 <b>Дата:</b> {date_str}\n"
                f"📦 <b>Товар:</b> {item_name}\n"
                f"🔖 <b>Артикул:</b> {nm_id_link}\n"
                f"⭐ <b>Рейтинг:</b> {rating}\n"
                f"💬 <b>Отзывы:</b> {reviews}\n"
                f"🚚 <b>Отгрузка:</b> {warehouse_name}\n"
                f"💰 <b>Логистика:</b> {delivery_rub:.2f}\n"
                f"🏙 <b>Доставка:</b> {region_name}\n"
                f"🛒 <b>Сегодня выкупов:</b> {today_count}\n"
                f"💲 <b>Сумма:</b> {base_price:.2f}  |  💸 <b>Комиссия:</b> {commision}%\n"
                f"🔽 <b>Цена с СПП:</b> {final_price:.2f}\n"
                f"📊 <b>Выкупов за 3 месяца:</b> {sales_last_3_months}\n"
                f"\n"
                f"📦 <b>Остаток:</b> {total_stocks} шт. ⏳ <b>Хватит примерно на:</b> {days_coverage:.0f} дн."
            )

            # Рассылаем всем пользователям
            for user in users:
                chat_id = user.telegram_id
                try:
                    if image_url:
                        try:
                            await bot.send_photo(
                                chat_id=chat_id,
                                photo=image_url,
                                caption=caption_text,
                                parse_mode="HTML"
                            )
                        except Exception:
                            fallback_text = f"{image_url}\n{caption_text}"
                            await bot.send_message(chat_id=chat_id, text=fallback_text, parse_mode="HTML")
                    else:
                        await bot.send_message(chat_id=chat_id, text=caption_text, parse_mode="HTML")

                except Exception as e:
                    print(f"Ошибка при отправке пользователю {chat_id}: {e}")

    session.close()
    print("Уведомления о новых выкупах отправлены!")

async def generate_daily_excel_report(token_id: int) -> bytes:
    """
    Формирует Excel-отчёт (в виде байтов) за последние сутки
    по конкретному токену (т.е. для конкретного пользователя).
    Включает следующие листы:
      1) Заказы
      2) Выкупы
      3) Отказы
      4) Товары с нулевыми остатками
    """

    session = SessionLocal()

    # Формируем период "последние сутки" (24 часа)
    now = datetime.datetime.utcnow()
    date_from = now - datetime.timedelta(days=1)

    # Подготовим Workbook
    wb = Workbook()
    ws_orders: Worksheet = wb.active
    ws_orders.title = "Orders"

    # 1) Лист с заказами (Order), у которых date >= date_from
    #    и привязан к нужному token_id
    orders = (
        session.query(Order)
        .filter(
            Order.token_id == token_id,
            Order.date >= date_from,
            Order.date <= now
        )
        .all()
    )

    # Заполним заголовки
    ws_orders.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион"])

    # Окрашиваем заголовки
    for cell in ws_orders[1]:
        if cell.column > 1:  # пропустим A1, оно у нас пустое
            cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")

    row_index = 2
    for o in orders:
        # В ячейки B..G запишем основные данные
        # А (колонка A) — под картинку

        date_val = o.date.strftime("%Y-%m-%d %H:%M:%S") if o.date else ""
        nm_id = o.nm_id or ""
        product_name = o.subject or ""
        price_val = o.price_with_disc or 0
        warehouse_val = o.warehouse_name or ""
        region_val = o.region_name or ""

        # Пишем: (B, C, D, E, F, G) => (date_val, srid, product_name, price, warehouse, region)
        ws_orders.cell(row=row_index, column=2, value=date_val)      # B
        ws_orders.cell(row=row_index, column=3, value=nm_id)      # C
        ws_orders.cell(row=row_index, column=4, value=product_name)  # D
        ws_orders.cell(row=row_index, column=5, value=price_val)     # E
        ws_orders.cell(row=row_index, column=6, value=warehouse_val) # F
        ws_orders.cell(row=row_index, column=7, value=region_val)    # G

        # Достаём product для вставки картинки (если есть)
        product = session.query(Product).filter_by(nm_id=o.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                # Открываем как PIL image
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                # Приводим к 80×80
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)

                # Превращаем обратно в BytesIO
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                # Создаём объект рисунка для openpyxl
                excel_img = ExcelImage(new_img_bytes)
                # Добавляем на лист, в ячейку A{row_index}
                cell_position = f"A{row_index}"
                ws_orders.add_image(excel_img, cell_position)

                # Поднимем высоту строки
                # 60 points ~ 80 px, можно ещё увеличить
                ws_orders.row_dimensions[row_index].height = 60

            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={o.nm_id}: {exc}")

        row_index += 1

    # Зададим одинаковую ширину для всех используемых колонок (A..G)
    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ws_orders.column_dimensions[col_letter].width = 15

    # 2) Лист с выкупами (Sale)
    ws_sales = wb.create_sheet(title="Выкупы")
    sales = (
        session.query(Sale)
        .filter(
            Sale.token_id == token_id,
            Sale.date >= date_from,
            Sale.date <= now
        )
        .all()
    )
    ws_sales.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион"])

    # окрашиваем заголовки
    for cell in ws_sales[1]:
        if cell.column > 1:  # пропустим A1, оно у нас пустое
            cell.fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")

    row_index = 2
    for s in sales:
        # В ячейки B..G запишем основные данные
        # А (колонка A) — под картинку

        date_val = s.date.strftime("%Y-%m-%d %H:%M:%S") if s.date else ""
        nm_id = s.nm_id or ""
        product_name = s.subject or ""
        price_val = s.price_with_disc or 0
        warehouse_val = s.warehouse_name or ""
        region_val = s.region_name or ""

        # Пишем: (B, C, D, E, F, G) => (date_val, srid, product_name, price, warehouse, region)
        ws_sales.cell(row=row_index, column=2, value=date_val)      # B
        ws_sales.cell(row=row_index, column=3, value=nm_id)      # C
        ws_sales.cell(row=row_index, column=4, value=product_name)  # D
        ws_sales.cell(row=row_index, column=5, value=price_val)     # E
        ws_sales.cell(row=row_index, column=6, value=warehouse_val) # F
        ws_sales.cell(row=row_index, column=7, value=region_val)    # G

        # Достаём product для вставки картинки (если есть)
        product = session.query(Product).filter_by(nm_id=s.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                # Открываем как PIL image
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                # Приводим к 80×80
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)

                # Превращаем обратно в BytesIO
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                # Создаём объект рисунка для openpyxl
                excel_img = ExcelImage(new_img_bytes)
                # Добавляем на лист, в ячейку A{row_index}
                cell_position = f"A{row_index}"
                ws_sales.add_image(excel_img, cell_position)

                # Поднимем высоту строки
                # 60 points ~ 80 px, можно ещё увеличить
                ws_sales.row_dimensions[row_index].height = 60

            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={o.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ws_sales.column_dimensions[col_letter].width = 15

    # 3) Лист с отказами (Order, is_cancel = True)
    ws_cancels = wb.create_sheet(title="Отказы")
    cancels = (
        session.query(Order)
        .filter(
            Order.token_id == token_id,
            Order.is_cancel == True,
            Order.date >= date_from,
            Order.date <= now
        )
        .all()
    )
    ws_cancels.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион", "Отменён?"])

    # окрашиваем заголовки
    for cell in ws_cancels[1]:
        if cell.column > 1:
            cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    row_index = 2
    for c in cancels:
        date_val = c.date.strftime("%Y-%m-%d %H:%M:%S") if c.date else ""
        nm_id = c.nm_id or ""
        product_name = c.subject or ""
        price_val = c.price_with_disc or 0
        warehouse_val = c.warehouse_name or ""
        region_val = c.region_name or ""
        is_cancel_val = "Да" if c.is_cancel else "Нет"

        # Пишем: (B, C, D, E, F, G) => (date_val, srid, product_name, price, warehouse, region)
        ws_cancels.cell(row=row_index, column=2, value=date_val)      # B
        ws_cancels.cell(row=row_index, column=3, value=nm_id)      # C
        ws_cancels.cell(row=row_index, column=4, value=product_name)  # D
        ws_cancels.cell(row=row_index, column=5, value=price_val)     # E
        ws_cancels.cell(row=row_index, column=6, value=warehouse_val) # F
        ws_cancels.cell(row=row_index, column=7, value=region_val)    # G
        ws_cancels.cell(row=row_index, column=8, value=is_cancel_val) # H


        # Достаём product для вставки картинки (если есть)
        product = session.query(Product).filter_by(nm_id=c.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                # Открываем как PIL image
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                # Приводим к 80×80
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)

                # Превращаем обратно в BytesIO
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                # Создаём объект рисунка для openpyxl
                excel_img = ExcelImage(new_img_bytes)
                # Добавляем на лист, в ячейку A{row_index}
                cell_position = f"A{row_index}"
                ws_cancels.add_image(excel_img, cell_position)

                # Поднимем высоту строки
                # 60 points ~ 80 px, можно ещё увеличить
                ws_cancels.row_dimensions[row_index].height = 60

            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={o.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 9):
        col_letter = get_column_letter(col)
        ws_cancels.column_dimensions[col_letter].width = 15

    # 4) Товары с нулевыми остатками (Stock, quantity = 0).
    #    Или, если нужно: те товары, которых вообще нет на складах
    #    (т.е. отсутствуют записи в Stock). Здесь вариант с quantity=0.
    ws_out_of_stock = wb.create_sheet(title="Отсутствие товаров")
    zero_stocks = (
        session.query(Stock)
        .filter(
            Stock.token_id == token_id,
            Stock.quantity == 0
        )
        .all()
    )
    ws_out_of_stock.append(["", "Артикул", "Склад", "Остаток", "Дата обновления"])

    # окрашиваем заголовки
    for cell in ws_out_of_stock[1]:
        cell.fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")

    for z in zero_stocks:
        nm_id = z.nm_id or ""
        warehouse_val = z.warehouseName or ""
        quantity_val = z.quantity or 0
        date_val = z.last_change_date.strftime("%Y-%m-%d %H:%M:%S") if z.last_change_date else ""

        # Пишем: (B, C, D, E, F, G) => (date_val, srid, product_name, price, warehouse, region)
        ws_out_of_stock.cell(row=row_index, column=2, value=nm_id)         # B
        ws_out_of_stock.cell(row=row_index, column=3, value=warehouse_val) # C
        ws_out_of_stock.cell(row=row_index, column=4, value=quantity_val)  # D
        ws_out_of_stock.cell(row=row_index, column=5, value=date_val)      # E

        # Достаём product для вставки картинки (если есть)
        product = session.query(Product).filter_by(nm_id=z.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                # Открываем как PIL image
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                # Приводим к 80×80
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)

                # Превращаем обратно в BytesIO
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                # Создаём объект рисунка для openpyxl
                excel_img = ExcelImage(new_img_bytes)
                # Добавляем на лист, в ячейку A{row_index}
                cell_position = f"A{row_index}"
                ws_out_of_stock.add_image(excel_img, cell_position)

                # Поднимем высоту строки
                # 60 points ~ 80 px, можно ещё увеличить
                ws_out_of_stock.row_dimensions[row_index].height = 60

            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={o.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 6):
        col_letter = get_column_letter(col)
        ws_out_of_stock.column_dimensions[col_letter].width = 15

    # Границы для всех листов
    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )

    # Выравнивание по центру
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Функция для применения стилей ко всем ячейкам
    def apply_styles_to_worksheet(ws):
        # Для всех занятых строк и столбцов
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment


    # Применяем стили ко всем листам
    for sheet in [ws_orders, ws_sales, ws_cancels, ws_out_of_stock]:
        apply_styles_to_worksheet(sheet)

    session.close()

    # Преобразуем Workbook в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

async def send_daily_reports_to_all_users(bot: Bot):
    """
    Функция, которую вызовет шедулер (либо вручную) раз в день.
    Проходимся по всем пользователям, у которых есть token_id,
    генерируем Excel и отправляем им в личку.
    """
    print("Отправляем ежедневные отчёты всем пользователям...")
    session = SessionLocal()
    users = (
        session.query(User)
        .filter(User.token_id.isnot(None))
        .filter(User.notify_daily_report == True)
        .all()
        )

    for user in users:
        token_id = user.token_id
        telegram_id = user.telegram_id

        # Генерируем Excel-отчёт (в виде байтов)
        report_bytes = await generate_daily_excel_report(token_id)

        # Длина файла
        file_size = len(report_bytes)
        # Если нужно, проверяем, не больше ли 50MB (лимит Телеграма)
        if file_size > 50_000_000:
            # Предположим, просто отправим сообщение об ошибке:
            await bot.send_message(chat_id=telegram_id, text="Отчёт слишком большой.")
            continue

        # Формируем документ
        doc = BufferedInputFile(file=report_bytes, filename="daily_report.xlsx")
        # Отправляем пользователю
        caption_text = "Ежедневный отчёт за последние 24 часа"
        await bot.send_document(chat_id=telegram_id, document=doc, caption=caption_text)

    session.close()

async def notify_subscription_expiring(bot: Bot):
    """
    Ищет токены, у которых subscription_until < now + WARNING_DAYS_LEFT,
    отправляет уведомление всем связанным пользователям.
    """
    WARNING_DAYS_LEFT = 3

    session = SessionLocal()

    now_utc = datetime.datetime.utcnow()
    warn_deadline = now_utc + timedelta(days=WARNING_DAYS_LEFT)

    # Ищем все токены, у которых подписка не истекла, но закончится в ближайшие WARNING_DAYS_LEFT
    # Не берем токены, у которых subscription_until=None или дата в прошлом (уже истекла)
    tokens_expiring = (
        session.query(Token)
        .filter(Token.subscription_until != None)            # есть дата
        .filter(Token.subscription_until > now_utc)         # ещё не истекла
        .filter(Token.subscription_until <= warn_deadline)  # но закончится в ближайшее время
        .all()
    )

    if not tokens_expiring:
        session.close()
        return  # ничего не делаем

    for token_obj in tokens_expiring:
        # Считаем, сколько осталось дней
        days_left = (token_obj.subscription_until - now_utc).days
        if days_left < 0:
            # уже истекло, пропускаем
            continue

        # Ищем всех пользователей, кто использует этот token
        users = session.query(User).filter_by(token_id=token_obj.id).all()
        if not users:
            continue

        # Формируем текст уведомления
        role_str = token_obj.role or "free"
        # Например:
        text = (
            f"Ваша подписка (<b>{role_str}</b>) истекает через <b>{days_left} дн.</b>\n"
            f"Дата окончания: {token_obj.subscription_until.strftime('%Y-%m-%d %H:%M:%S')}\n"
            "Чтобы продлить подписку, воспользуйтесь разделом оплаты."
        )

        # Рассылаем всем пользователям
        for user in users:
            chat_id = user.telegram_id
            if chat_id:
                try:
                    await bot.send_message(chat_id=chat_id, text=text, parse_mode="HTML")
                except Exception as e:
                    print(f"Ошибка при отправке уведомления пользователю {chat_id}: {e}")

    session.close()