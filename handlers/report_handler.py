import io
from datetime import datetime, timedelta

from aiogram import types, Dispatcher
from aiogram.filters import Command
from aiogram.types import BufferedInputFile
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, PieChart, BarChart3D
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.plotarea import PlotArea
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import GradientFillProperties, GradientStop
from PIL import Image as PILImage
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList

from sqlalchemy import func, desc,and_, case

from db.database import SessionLocal
from db.models import Product, Order, Sale, Stock, User

from core.sub import user_has_role

MAX_TELEGRAM_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

async def cmd_my_products(message: types.Message, user_id:int = None, days: int = 0):
    """Хендлер, который генерирует полный Excel-отчёт с двумя листами:
       1) Группировка по категориям
       2) Топ по заказам/выкупам за 7 дней
    """

    if user_id is None:
        user_id = message.from_user.id

    # 1) Если days == 0, значит вызываем /my_products напрямую
    #    Показываем inline-кнопки
    if days == 0:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="7 дней", callback_data="my_products_7")],
                [InlineKeyboardButton(text="30 дней", callback_data="my_products_30")],
                [InlineKeyboardButton(text="90 дней", callback_data="my_products_90")]
            ]
        )
        await message.answer("За какой период вывести отчёт?", reply_markup=kb)
        return  # Прерываем, отчёт пока не генерируем

    # 2) Иначе, days != 0: значит, попали сюда из коллбэка
    #    Генерируем отчёт
    session = SessionLocal()
    db_user = session.query(User).filter_by(telegram_id=str(user_id)).first()
    if not db_user or not db_user.token_id:
        await message.answer("Нет привязанного токена. Сначала /start и пришлите токен.")
        session.close()
        return

    if days == 7:
        allowed_roles = ["base", "advanced", "test", "super"]
    elif days in (30, 90):
        allowed_roles = ["advanced", "test", "super"]
    else:
        # Если пользователь ввёл другие числа, то, например, только super
        allowed_roles = ["super"]

    # 3) Проверяем роль через user_has_role
    has_access = user_has_role(session, str(user_id), allowed_roles)
    if not has_access:
        await message.answer(
            f"У вас нет доступа к просмотру заказов за {days} дней.\n"
            f"Доступны только роли: {', '.join(allowed_roles)}."
        )
        session.close()
        return

    token_id = db_user.token_id

    products = session.query(Product).filter_by(token_id=token_id).all()
    wb = Workbook()

    generate_excel_grouped_by_subject(products, wb=wb, token_id=token_id)
    add_top_two_columns_sheet(wb, session, token_id=token_id, days=days, top_n=50)

    date_from = datetime.utcnow() - timedelta(days=days)
    date_to = datetime.utcnow()
    generate_detailed_sheets_for_products(session, wb, token_id, date_from, date_to)

    session.close()

    # Сериализация
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    workbook_bytes = output.getvalue()

    if len(workbook_bytes) > MAX_TELEGRAM_FILE_SIZE:
        await message.answer("Извините, файл слишком большой для отправки через Telegram!")
        return

    doc = BufferedInputFile(workbook_bytes, filename=f"сводный отчёт за {days} дней.xlsx")
    await message.answer_document(document=doc, caption=f"Ваш отчёт за {days} дней")

def generate_excel_grouped_by_subject(products: list[Product], wb: Workbook, token_id) -> None:
    """
    Создаёт / заполняет первый лист в переданной книге wb:
    - Группирует товары по subject_name
    - Горизонтальный вывод (ID + картинки)
    """
    from collections import defaultdict

    ws = wb.active
    ws.title = "Сводка по категориям"

    # Фильтруем товары по token_id
    products = [p for p in products if p.token_id == token_id]

    # Группируем товары
    grouped = defaultdict(list)
    for p in products:
        grouped[p.subject_name].append(p)

    sorted_subjects = sorted(grouped.keys())

    # Стили для категории
    category_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
    category_font = Font(bold=True, color="FFFFFFFF")
    category_align = Alignment(horizontal="center", vertical="center")

    # 1) Определим, сколько максимум столбцов нам может понадобиться
    max_columns = 0  
    for subject_name in sorted_subjects:
        num_products = len(grouped[subject_name])
        if num_products > max_columns:
            max_columns = num_products

    current_row = 1
    for subject_name in sorted_subjects:
        products_in_group = grouped[subject_name]
        num_products = len(products_in_group)
        if num_products == 0:
            continue

        # --- 2) Вывод названия категории ---
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=num_products
        )
        cell = ws.cell(row=current_row, column=1, value=subject_name)
        cell.fill = category_fill
        cell.font = category_font
        cell.alignment = category_align
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # --- 3 Строка с ID товаров ---
        row_id = current_row
        for col_idx, product in enumerate(products_in_group, start=1):
            c = ws.cell(row=row_id, column=col_idx)
            c.value = f"ID:{product.nm_id}"
            c.alignment = Alignment(horizontal="center")

        current_row += 1

        # --- 4) Строка с картинками ---
        images_row = current_row
        for col_idx, product in enumerate(products_in_group, start=1):
            col_letter = get_column_letter(col_idx)
            if product.resize_img:
                try:
                    img_bytes = io.BytesIO(product.resize_img)
                    pil_img = PILImage.open(img_bytes)
                    pil_img = pil_img.resize((152, 200), PILImage.Resampling.LANCZOS)
                    img_bytes = io.BytesIO()
                    pil_img = pil_img.convert("RGB")
                    pil_img.save(img_bytes, format="JPEG", optimize=True, quality=70)
                    img_bytes.seek(0)

                    excel_img = ExcelImage(img_bytes)
                    ws.add_image(excel_img, f"{col_letter}{images_row}")
                except Exception as e:
                    print(f"Ошибка при загрузке картинки для nm_id={product.nm_id}: {e}")

        # Задаём высоту строки для картинок (в пунктах)
        ws.row_dimensions[images_row].height = 150
        current_row += 1

    # --- 5) Теперь, когда мы знаем, что max_columns — это максимум занятых столбцов,
    #        пройдёмся и выставим ширину 21.7 (или сколько вы подобрали) всем им
    for col_idx in range(1, max_columns + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 21.7

def add_top_two_columns_sheet(wb: Workbook, session, token_id:int, days: int = 90, top_n=50):
    """
    Создаёт в книге wb лист со стилем а-ля «две колонки»: слева ТОП заказов, справа ТОП выкупов.
    Параметры:
      - days: за какой период берём (7, 30, 90 и т.д.)
      - top_n: сколько позиций брать в ТОП (по умолчанию 3)
      - token_id: для фильтрации товаров
    """

    ws = wb.create_sheet(title="Топ отчёт")

    # -- 1. Шапка (строка 2) --
    # Сольём колонки A..C для "ТОП ЗАКАЗЫ"
    ws.merge_cells("A2:C2")
    top_orders_cell = ws["A2"]
    top_orders_cell.value = "ТОП ЗАКАЗЫ"
    top_orders_cell.fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")  
    top_orders_cell.font = Font(bold=True, color="FFFFFFFF")
    top_orders_cell.alignment = Alignment(horizontal="center")

    # Сольём колонки D..F для "ТОП ВЫКУПЫ"
    ws.merge_cells("D2:F2")
    top_sales_cell = ws["D2"]
    top_sales_cell.value = "ТОП ВЫКУПЫ"
    top_sales_cell.fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")  
    top_sales_cell.font = Font(bold=True, color="FFFFFFFF")
    top_sales_cell.alignment = Alignment(horizontal="center")

    # Можно подогнать высоту строки 2
    ws.row_dimensions[2].height = 25

    # -- 2. Получаем данные TOP N заказов и TOP N выкупов за days --

    date_from = datetime.utcnow() - timedelta(days=days)

    # 2.1 Топ заказов (Order):
    # Считаем:
    #   count(*)  - сколько заказов
    #   sum(price_with_disc) - общая сумма
    #   count(*) с is_cancel=True - сколько отказов
    #   sum(price_with_disc) там же - сумма отказов
    # Группируем по nm_id и сортируем по count(*) (самый популярный товар)
    top_orders_raw = (
        session.query(
            Order.nm_id,
            func.count(Order.id).label("cnt_orders"),
            func.sum(Order.price_with_disc).label("sum_orders"),
             func.sum(case((Order.is_cancel == True, 1),else_=0)).label("cnt_cancel"),
             func.sum(case((Order.is_cancel == True, Order.price_with_disc),else_=0)).label("sum_cancel"),
        )
        .filter(and_(Order.date >= date_from, Order.date != None, Order.token_id == token_id))
        .group_by(Order.nm_id)
        .order_by(desc("cnt_orders"))
        .limit(top_n)
        .all()
    )

    # Превратим в удобный список dict и дотянем название/картинку из products
    top_orders = []
    for row in top_orders_raw:
        nm_id = row.nm_id
        product = session.query(Product).filter_by(nm_id=nm_id).first()
        title = product.supplier_article if product else f"Товар {nm_id}"
        image_url = product.image_url if product else None
        top_orders.append({
            "nm_id": nm_id,
            "title": title,
            "count": int(row.cnt_orders or 0),
            "sum": float(row.sum_orders or 0.0),
            "count_cancel": int(row.cnt_cancel or 0),
            "sum_cancel": float(row.sum_cancel or 0.0),
            "image_url": image_url,
            "resize_img": product.resize_img if product else None
        })

    # 2.2 Топ выкупов (Sale):
    # Аналогично, но учитываем "возврат" = sale_id starts with 'R'
    top_sales_raw = (
        session.query(
            Sale.nm_id,
            func.count(Sale.id).label("cnt_sales"),
            func.sum(Sale.price_with_disc).label("sum_sales"),
            func.sum(case((Sale.sale_id.like("R%"), 1),else_=0)).label("cnt_return"),
            func.sum(case((Sale.sale_id.like("R%"), Sale.price_with_disc), else_=0)).label("sum_return"),)
        .filter(and_(Sale.date >= date_from, Sale.date != None, Sale.token_id == token_id))
        .group_by(Sale.nm_id)
        .order_by(desc("cnt_sales"))
        .limit(top_n)
        .all()
    )

    top_sales = []
    for row in top_sales_raw:
        nm_id = row.nm_id
        product = session.query(Product).filter_by(nm_id=nm_id).first()
        title = product.supplier_article if product else f"Товар {nm_id}"
        image_url = product.image_url if product else None
        top_sales.append({
            "nm_id": nm_id,
            "title": title,
            "count": int(row.cnt_sales or 0),
            "sum": float(row.sum_sales or 0.0),
            "count_return": int(row.cnt_return or 0),
            "sum_return": float(row.sum_return or 0.0),
            "image_url": image_url,
            "resize_img": product.resize_img if product else None
        })

    # -- 3. Записываем в виде блоков: слева топ заказов, справа топ выкупов --
    #   Каждый элемент займёт ~6-7 строк. Картинку поставим в первую часть, текст ниже.

    # Начинаем с ~ 3 строки (под шапкой).
    row_start = 3

    # Предлагаем взять max из обоих списков (на случай, если в одном 3 товара, а в другом меньше)
    max_len = max(len(top_orders), len(top_sales))

    # Подготовим стили (например, чёрный фон, оранжевый текст и т.п.)
    black_fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    orange_font = Font(color="FFFFA500", bold=False)
    white_font = Font(color="FFFFFFFF", bold=False)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Зададим ширины столбцов
    ws.column_dimensions["A"].width = 2   # чуть узкая, «поле»
    ws.column_dimensions["B"].width = 21.7  # под картинку
    ws.column_dimensions["C"].width = 30  # под текст
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 21.7
    ws.column_dimensions["F"].width = 30

    for i in range(max_len):
        order_item = top_orders[i] if i < len(top_orders) else None
        sale_item = top_sales[i] if i < len(top_sales) else None

        # -- Левый блок (заказы) в колонках B..C
        if order_item:
            # 3.1 Картинка (вставим в B.. - фактически можно просто B=row_start)
            if order_item["resize_img"]:
                try:
                    img_bytes = io.BytesIO(order_item["resize_img"])
                    pil_img = PILImage.open(img_bytes)
                    # Масштабируем, скажем, до 150x150
                    pil_img = pil_img.resize((152,200), PILImage.Resampling.LANCZOS)
                    img_bytes = io.BytesIO()
                    pil_img = pil_img.convert("RGB")
                    pil_img.save(img_bytes, format="JPEG", optimize=True, quality=70)
                    img_bytes.seek(0)
                    excel_img = ExcelImage(img_bytes)
                    # Вставляем в B{row_start}, Excel сам растянет
                    ws.add_image(excel_img, f"B{row_start}")
                    ws.row_dimensions[row_start].height = 150
                except:
                    pass

            # 3.2 Текст (C row_start + ниже)
            # Можно собрать текст заказы
            text_lines = []
            text_lines.append(order_item["title"])
            text_lines.append(f"WB Страница")
            text_lines.append(f"Заказано: {order_item['count']} шт.")
            text_lines.append(f"На сумму: {int(order_item['sum'])} руб.")
            text_lines.append(f"Из них отказов: {order_item['count_cancel']} шт.")
            text_lines.append(f"На сумму: {int(order_item['sum_cancel'])} руб.")

            cell_text = "\n".join(text_lines)
            text_cell = ws.cell(row=row_start, column=3, value=cell_text)
            text_cell.fill = black_fill
            text_cell.font = white_font
            text_cell.alignment = left_align

        # -- Правый блок (выкупы) в колонках E..F
        if sale_item:
            if sale_item["resize_img"]:
                try:
                    img_bytes = io.BytesIO(sale_item["resize_img"])
                    pil_img = PILImage.open(img_bytes)
                    pil_img = pil_img.resize((152,200), PILImage.Resampling.LANCZOS)
                    img_bytes = io.BytesIO()
                    pil_img = pil_img.convert("RGB")
                    pil_img.save(img_bytes, format="JPEG", optimize=True, quality=70)
                    img_bytes.seek(0)
                    excel_img = ExcelImage(img_bytes)
                    ws.add_image(excel_img, f"E{row_start}")
                except:
                    pass

            text_lines = []
            text_lines.append(sale_item["title"])
            text_lines.append(f"WB Страница")
            text_lines.append(f"Выкуплено: {sale_item['count']} шт.")
            text_lines.append(f"На сумму: {int(sale_item['sum'])} руб.")
            text_lines.append(f"Из них возвратов: {sale_item['count_return']} шт.")
            text_lines.append(f"На сумму: {int(sale_item['sum_return'])} руб.")

            text_cell = ws.cell(row=row_start, column=6, value="\n".join(text_lines))
            text_cell.fill = black_fill
            text_cell.font = white_font
            text_cell.alignment = left_align

        # Каждый «блок» займёт 6-7 строк в высоту, чтобы картинка + текст влезли
        # Для простоты можно не использовать merged cells, а просто «прыгать» вниз.
        # Скажем, прыгаем на 6 строк.
        row_start += 6

      # 3.2 Заливаем колонку D любым цветом (например, светло-серым "FFD3D3D3")
    separator_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    max_row = ws.max_row  # сколько строк реально получилось
    for r in range(1, max_row + 1):
        ws.cell(row=r, column=4).fill = separator_fill

def generate_detailed_sheets_for_products(session, wb: Workbook, token_id:int, date_from: datetime, date_to: datetime):
    """
    Для каждого товара создаём отдельный лист, в котором:
      - Зеленая шапка с названием товара и периодом
      - В A3 картинка (170x200)
      - В B3 - заказы, сумма, отказы (чёрный фон, белый текст, \n внутри)
      - В C3 - выкупы, сумма, возвраты (чёрный фон, белый текст, \n внутри)
    """
    # 1) Получаем товары (можно фильтровать только нужные)
    products = session.query(Product).filter_by(token_id=token_id).all()

    for product in products:

        # Создаём основной лист
        nm_id = product.nm_id
        sheet_name = str(nm_id)[:31]  # Excel не любит имена >31 символ
        ws = wb.create_sheet(title=sheet_name)

        # Создаём технический лист
        data_sheet_name = f"data{nm_id}"
        data_ws = wb.create_sheet(title=data_sheet_name[:31])  # тоже обрежем до 31 символа

        # 2) Зелёная шапка (A1..C1)S
        ws.merge_cells("A1:C1")
        cell_title = ws["A1"]
        # Пишем и товар, и период
        cell_title.value = (
            f"Детальная статистика по товару: {product.supplier_article}\n"
            f"(Период: {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')})"
        )
        cell_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Заливка и шрифт
        cell_title.fill = PatternFill(start_color="FF008000", end_color="FF008000", fill_type="solid")  # ярко-зелёный
        cell_title.font = Font(bold=True, color="FFFFFFFF")

        # Увеличим высоту для шапки
        ws.row_dimensions[1].height = 30

        # 3) В A2 вставим картинку
        if product.resize_img:
            try:
                img_bytes = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_bytes)
                # Масштабируем под 152x200
                pil_img = pil_img.resize((152, 200), PILImage.Resampling.LANCZOS)

                img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(img_bytes, format="JPEG", optimize=True, quality=70)
                img_bytes.seek(0)

                excel_img = ExcelImage(img_bytes)
                ws.add_image(excel_img, "A2")

                # Поднимаем высоту строки 3 (200 пунктов ~ 266 пикселей, но можно подобрать)
                ws.row_dimensions[2].height = 150  # или 180, 200, экспериментируйте
            except Exception as e:
                print(f"Не удалось загрузить картинку для nm_id={nm_id}: {e}")

        # 4) Считаем статистику по заказам (orders) за период
        total_orders = session.query(func.count(Order.id),func.sum(Order.price_with_disc),
    # вместо func.case
        func.sum(case((Order.is_cancel == True, 1),else_=0)).label("cnt_cancel"),func.sum(case((Order.is_cancel == True, Order.price_with_disc),else_=0)).label("sum_cancel"),).filter(
    Order.nm_id == nm_id,
    Order.date >= date_from,
    Order.date <= date_to,
    Order.token_id == token_id
).first()

        orders_count = total_orders[0] or 0
        orders_sum = total_orders[1] or 0
        orders_cnt_cancel = total_orders[2] or 0
        orders_sum_cancel = total_orders[3] or 0

        # 5) Считаем статистику по выкупам (sales) за период
        total_sales = (
    session.query(
        func.count(Sale.id),
        func.sum(Sale.price_with_disc),
        func.sum(
            case(
                (Sale.sale_id.like("R%"), 1),
                else_=0
            )
        ).label("cnt_return"),
        func.sum(
            case(
                (Sale.sale_id.like("R%"), Sale.price_with_disc),
                else_=0
            )
        ).label("sum_return"),
    )
    .filter(
        Sale.nm_id == nm_id,
        Sale.date >= date_from,
        Sale.date <= date_to,
        Sale.token_id == token_id
    )
    .first()
)

        sales_count = total_sales[0] or 0
        sales_sum = total_sales[1] or 0
        sales_cnt_return = total_sales[2] or 0
        sales_sum_return = total_sales[3] or 0

        # 6) Выводим всё в B3 (заказы) и C3 (выкупы), чёрный фон, белый текст
        black_fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
        white_font = Font(color="FFFFFFFF")
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

        orders_text_lines = [
            f"Заказов: {orders_count} шт.",
            f"На сумму: {int(orders_sum)} руб.\n\n\n\n\n\n",
            f"Отказов: {int(orders_cnt_cancel)} шт.",
            f"На сумму: {int(orders_sum_cancel)} руб.",
        ]
        orders_text = "\n".join(orders_text_lines)
        b3 = ws.cell(row=2, column=2, value=orders_text)
        b3.fill = black_fill
        b3.font = white_font
        b3.alignment = left_align

        sales_text_lines = [
            f"Выкуплено: {sales_count} шт.",
            f"На сумму: {int(sales_sum)} руб.\n\n\n\n\n\n",
            f"Возвратов: {int(sales_cnt_return)} шт.",
            f"На сумму: {int(sales_sum_return)} руб.",
        ]
        sales_text = "\n".join(sales_text_lines)
        c3 = ws.cell(row=2, column=3, value=sales_text)
        c3.fill = black_fill
        c3.font = white_font
        c3.alignment = left_align

        # 7) Подогнать ширину столбцов
        ws.column_dimensions["A"].width = 21.7
        ws.column_dimensions["B"].width = 30    
        ws.column_dimensions["C"].width = 30

        date_list = []
        cur_date = date_from
        while cur_date <= date_to:
                date_list.append(cur_date)
                cur_date += timedelta(days=1)

        daily_stats = []
        for d in date_list:
            day_orders_count = session.query(func.count(Order.id)) \
                .filter(Order.nm_id == nm_id,  func.DATE(Order.date) == d.date()) \
                .scalar() or 0
            day_cancel_count = session.query(func.count(Order.id)) \
                .filter(Order.nm_id == nm_id, Order.is_cancel == True,
                            func.DATE(Order.date) == d.date()) \
                .scalar() or 0
            day_sales_count = session.query(func.count(Sale.id)) \
                .filter(Sale.nm_id == nm_id, func.DATE(Sale.date) == d.date()) \
                .scalar() or 0
            day_return_count = session.query(func.count(Sale.id)) \
                .filter(Sale.nm_id == nm_id, Sale.sale_id.like("R%"),
                            func.DATE(Sale.date) == d.date()) \
                .scalar() or 0

            daily_stats.append((d.date(), day_orders_count, day_cancel_count, day_sales_count, day_return_count))

        # Заказы с регионов

        region_stats = (
            session.query(
                Order.region_name,
                func.count(Order.id).label("cnt_orders")
            )
            .filter(
                Order.nm_id == nm_id,
                Order.date >= date_from,
                Order.date <= date_to,
                Order.is_cancel == False  # если не хотим считать «отменённые»
            )
            .group_by(Order.region_name)
            .order_by(desc("cnt_orders"))
            .all()
        )

        # Заказы со складов
        warehouse_stats = (
            session.query(
                Order.warehouse_name,
                func.count(Order.id).label("cnt_orders")
            )
            .filter(
                Order.nm_id == nm_id,
                Order.date >= date_from,
                Order.date <= date_to,
                Order.is_cancel == False  # Исключаем отменённые заказы
            )
            .group_by(Order.warehouse_name)
            .order_by(desc("cnt_orders"))
            .all()
        )

        # Извлекаем данные из таблицы Stock
        stock_stats = (
            session.query(
                Stock.warehouseName,
                func.sum(Stock.quantity).label("quantity"),
                func.sum(func.coalesce(Stock.inWayToClient, 0)).label("inWayToClient")
            )
            .filter(
                Stock.nm_id == nm_id  # Фильтруем по артикулу товара
            )
            .group_by(Stock.warehouseName)
            .order_by(desc("quantity"))
            .all()
        )

        # Считаем общее количество на всех складах
        total_stock = sum(row.quantity for row in stock_stats)
        in_way_total = sum(row.inWayToClient for row in stock_stats)


        sales_by_warehouse = (
            session.query(
                Order.warehouse_name,
                func.count(Order.id).label("sales_count")
            )
            .filter(
                Order.nm_id == nm_id,
                Order.date >= date_from,
                Order.date <= date_to,
                Order.is_cancel == False
            )
            .group_by(Order.warehouse_name)
            .all()
        )

        # Шаг 2. Собираем данные о текущих остатках
        stock_by_warehouse = {
            row.warehouseName: row.quantity
            for row in session.query(Stock.warehouseName, Stock.quantity)
            .filter(Stock.nm_id == nm_id)
        }

        # Шаг 3. Рассчитываем средний спрос и рекомендуемые запасы
        recommendations = []
        total_stock_available = sum(stock_by_warehouse.values())
        total_sales = sum(row.sales_count for row in sales_by_warehouse)
        reserve_stock = int(0.1 * total_stock_available)  # 10% от общего запаса в резерв

        for sale in sales_by_warehouse:
            warehouse = sale.warehouse_name
            avg_demand = int(sale.sales_count)
            current_stock = int(stock_by_warehouse.get(warehouse, 0))
            recommended_stock = int(avg_demand + (reserve_stock / len(sales_by_warehouse)))
            difference = int(recommended_stock - current_stock)
            recommendations.append((warehouse, avg_demand, current_stock, recommended_stock, difference))

# region_stats вернёт список кортежей [(region1, count1), (region2, count2), ...]

            # (5) Записываем таблицу для 1 графика
        start_row = 1
        data_ws.cell(row=start_row, column=1, value="Дата")
        data_ws.cell(row=start_row, column=2, value="Заказы")
        data_ws.cell(row=start_row, column=3, value="Отказы")
        data_ws.cell(row=start_row, column=4, value="Выкупы")
        data_ws.cell(row=start_row, column=5, value="Возвраты")

        row = start_row + 1
        for (dt, orders_cnt, cancel_cnt, sales_cnt, return_cnt) in daily_stats:
            data_ws.cell(row=row, column=1, value=dt.strftime("%d.%m"))
            data_ws.cell(row=row, column=2, value=orders_cnt)
            data_ws.cell(row=row, column=3, value=cancel_cnt)
            data_ws.cell(row=row, column=4, value=sales_cnt)
            data_ws.cell(row=row, column=5, value=return_cnt)
            row += 1

        data_end = row - 1  # последняя строка с данными

        # Записываем таблицу для 2 графика

        region_start_row = data_end + 2  # немного отступаем от предыдущих записей

        data_ws.cell(row=region_start_row, column=1, value="Регион")
        data_ws.cell(row=region_start_row, column=2, value="Кол-во заказов")

        r = region_start_row + 1
        for (region_name, cnt) in region_stats:
            combined_label = f"{region_name} -\n ({cnt} шт)"
            data_ws.cell(row=r, column=1, value=combined_label)
            data_ws.cell(row=r, column=2, value=cnt)
            r += 1

        region_data_end = r - 1

        # Записываем таблицу для 3 графика (Склады)
        warehouse_start_row = region_data_end + 2  # отступаем от предыдущих данных

        data_ws.cell(row=warehouse_start_row, column=1, value="Склад")
        data_ws.cell(row=warehouse_start_row, column=2, value="Кол-во заказов")

        r = warehouse_start_row + 1
        for (warehouse_name, cnt) in warehouse_stats:
            combined_label = f"{warehouse_name} -\n ({cnt} шт)"
            data_ws.cell(row=r, column=1, value=combined_label)
            data_ws.cell(row=r, column=2, value=cnt)
            r += 1

        warehouse_data_end = r - 1

        # Записываем таблицу для 4 графика (Склады)

        # Записываем таблицу для графика заполненности складов
        stock_start_row = warehouse_data_end + 2  # отступаем от предыдущих данных

        data_ws.cell(row=stock_start_row, column=1, value="Склад")
        data_ws.cell(row=stock_start_row, column=2, value="Количество на складе")

        r = stock_start_row + 1
        for (warehouse_name, total_quantity, _) in stock_stats:
            combined_label = f"{warehouse_name} -\n ({total_quantity} шт)"
            data_ws.cell(row=r, column=1, value=combined_label)
            data_ws.cell(row=r, column=2, value=total_quantity)
            r += 1

        stock_data_end = r - 1

        # Записываем таблицу для 5 графика (Рекомендации по заполнению складов)

        rec_start_row = stock_data_end + 2  # отступаем от предыдущих данных
        data_ws.cell(row=rec_start_row, column=1, value="Склад")
        data_ws.cell(row=rec_start_row, column=2, value="Средний спрос")
        data_ws.cell(row=rec_start_row, column=3, value="Текущий запас")
        data_ws.cell(row=rec_start_row, column=4, value="Рекомендованный запас")
        data_ws.cell(row=rec_start_row, column=5, value="Необходимое перемещение")

        row = rec_start_row + 1

        for warehouse, avg_demand, current_stock, recommended_stock, difference in recommendations:
            data_ws.cell(row=row, column=1, value=warehouse)
            data_ws.cell(row=row, column=2, value=avg_demand)
            data_ws.cell(row=row, column=3, value=current_stock)
            data_ws.cell(row=row, column=4, value=recommended_stock)
            data_ws.cell(row=row, column=5, value=difference)
            row += 1

        rec_data_end = row - 1
        if data_end > start_row:  # есть какие-то данные
            
            # (6) Создаём 1 график
            chart = LineChart()
            chart.title = "График продаж (№1)"
            chart.title.overlay = False
            # Категории = столбец A (даты)
            cats = Reference(data_ws, min_col=1, min_row=start_row+1, max_row=data_end)
            # Данные = столбцы B..E
            data = Reference(data_ws, min_col=2, max_col=5, min_row=start_row, max_row=data_end)
            # titles_from_data=True => берём заголовки из первой строки (row=start_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # Настраиваем серии (цвет, метки и т.д.)
            colors = ["7E57C2", "FF4343", "00B050", "FB9030"]    # градиент фиолетового (Material Palette).
            series_names = ["Заказы", "Отказы", "Выкупы", "Возвраты"]

            for i, s in enumerate(chart.series):
                # Вместо s.title = "Заказы" -> делаем SeriesLabel
                lbl = SeriesLabel()
                lbl.v = series_names[i]  # ваша строка, например "Заказы"
                s.tx = lbl  # присваиваем объект SeriesLabel
                s.graphicalProperties.line.solidFill = colors[i]
                s.marker.symbol = "circle"
                s.marker.size = 7
                s.marker.graphicalProperties.solidFill = colors[i]

            
            chart.y_axis.title = "Количество"
            chart.x_axis.title = "Дата"
            chart.legend.position = "r"

        
            chart.x_axis.delete = False
            chart.y_axis.delete = False

            # Проверяем chart.plot_area
            if chart.plot_area is None:
                chart.plot_area = PlotArea()

            if chart.plot_area.graphicalProperties is None:
                chart.plot_area.graphicalProperties = GraphicalProperties()

            if chart.graphical_properties is None:
                chart.graphical_properties = GraphicalProperties()


            chart.plot_area.graphicalProperties.noFill = True

            chart.graphical_properties.gradFill = GradientFillProperties(
                gsLst=(
                    GradientStop(pos=0, srgbClr="EEF1F6"),      # Начало градиента
                    GradientStop(pos=100000, srgbClr="C4CAE9"), # Конец градиента
                )
            )
            ws.add_chart(chart, "A3")
        if region_data_end > region_start_row:  # есть какие-то данные
            # Создаём 2 график

            pie = PieChart()
            pie.title = "Заказы по регионам"
            pie.title.overlay = False
            pie._3d = True
            pie.style = 26
            pie.legend.position = "l"
            # Категории — столбец A (region_name)
            cats = Reference(data_ws, min_col=1, min_row=region_start_row+1, max_row=region_data_end)
            # Данные — столбец B (cnt_orders)
            data = Reference(data_ws, min_col=2, min_row=region_start_row, max_row=region_data_end)

            pie.add_data(data, titles_from_data=True)  
            # В PieChart, если titles_from_data=True, он возьмёт "Кол-во заказов" 
            # как заголовок, а cats пойдёт отдельно:
            pie.set_categories(cats)

            # Настраиваем подписи данных
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showVal = True  # Показывать значения (количество заказов)
            pie.dataLabels.showCatName = False  # Показывать названия категорий (регионы)
            pie.dataLabels.showSerName = False
            pie.dataLabels.showPercent = False  # Показывать проценты
            pie.dataLabels.position = "inEnd"  # Метки внутри диаграммы
            pie.dataLabels.showLegendKey = False
            pie.dataLabels.showBubbleSize = False

            # Проверяем chart.plot_area
            if pie.plot_area is None:
                pie.plot_area = PlotArea()

            if pie.plot_area.graphicalProperties is None:
                pie.plot_area.graphicalProperties = GraphicalProperties()

            if pie.graphical_properties is None:
                pie.graphical_properties = GraphicalProperties()


            pie.plot_area.graphicalProperties.noFill = True

            pie.graphical_properties.gradFill = GradientFillProperties(
                gsLst=(
                    GradientStop(pos=0, srgbClr="EEF1F6"),      # Начало градиента
                    GradientStop(pos=100000, srgbClr="C4CAE9"), # Конец градиента
                )
            )

            pie.layout = Layout(
                manualLayout=ManualLayout(
                    x= 1,  # Сдвиг вправо на 70%
                )
            )
            ws.add_chart(pie, "D3")

        if warehouse_data_end > warehouse_start_row:  # есть какие-то данные
            # Создаём 3 график

            warehouse_pie = PieChart()
            warehouse_pie.title = "Заказы по складам"
            warehouse_pie.title.overlay = False
            warehouse_pie._3d = True
            warehouse_pie.style = 26
            warehouse_pie.legend.position = "l"

            # Категории — столбец A (warehouse_name)
            cats_warehouse = Reference(data_ws, min_col=1, min_row=warehouse_start_row+1, max_row=warehouse_data_end)
            # Данные — столбец B (cnt_orders)
            data_warehouse = Reference(data_ws, min_col=2, min_row=warehouse_start_row, max_row=warehouse_data_end)

            warehouse_pie.add_data(data_warehouse, titles_from_data=True)
            warehouse_pie.set_categories(cats_warehouse)

            # Настраиваем подписи данных
            warehouse_pie.dataLabels = DataLabelList()
            warehouse_pie.dataLabels.showVal = True  # Показывать значения (количество заказов)
            warehouse_pie.dataLabels.showCatName = False  # Показывать названия категорий (склады)
            warehouse_pie.dataLabels.showSerName = False
            warehouse_pie.dataLabels.showPercent = False  # Показывать проценты
            warehouse_pie.dataLabels.position = "inEnd"  # Метки внутри диаграммы
            warehouse_pie.dataLabels.showLegendKey = False
            warehouse_pie.dataLabels.showBubbleSize = False

            if warehouse_pie.plot_area is None:
                warehouse_pie.plot_area = PlotArea()

            if warehouse_pie.plot_area.graphicalProperties is None:
                warehouse_pie.plot_area.graphicalProperties = GraphicalProperties()

            if warehouse_pie.graphical_properties is None:
                warehouse_pie.graphical_properties = GraphicalProperties()

            # Настраиваем фон диаграммы
            warehouse_pie.plot_area.graphicalProperties.noFill = True
            warehouse_pie.graphical_properties.gradFill = GradientFillProperties(
                gsLst=(
                    GradientStop(pos=0, srgbClr="EEF1F6"),      
                    GradientStop(pos=100000, srgbClr="C4CAE9"), 
                )
            )

            # Сдвигаем диаграмму вправо
            warehouse_pie.layout = Layout(
                manualLayout=ManualLayout(
                    x=1,  # Сдвиг вправо
                )
            )
            ws.add_chart(warehouse_pie, "A17")

        if stock_data_end > stock_start_row:  # есть какие-то данные
            # Создаём 4 график

            stock_pie = PieChart()
            stock_pie.title = f"Заполненность складов (Всего: {total_stock} шт.\nВ пути: {in_way_total} шт.)"
            stock_pie.title.overlay = False
            stock_pie._3d = True
            stock_pie.style = 26
            stock_pie.legend.position = "l"

            # Категории — столбец A (warehouse_name)
            cats_stock = Reference(data_ws, min_col=1, min_row=stock_start_row+1, max_row=stock_data_end)
            # Данные — столбец B (total_quantity)
            data_stock = Reference(data_ws, min_col=2, min_row=stock_start_row, max_row=stock_data_end)

            stock_pie.add_data(data_stock, titles_from_data=True)
            stock_pie.set_categories(cats_stock)

            # Настраиваем подписи данных
            stock_pie.dataLabels = DataLabelList()
            stock_pie.dataLabels.showVal = True  # Показывать значения (количество товара)
            stock_pie.dataLabels.showCatName = False  # Показывать названия складов
            stock_pie.dataLabels.showSerName = False
            stock_pie.dataLabels.showPercent = False  # Показывать проценты
            stock_pie.dataLabels.position = "inEnd"  # Метки внутри диаграммы
            stock_pie.dataLabels.showLegendKey = False
            stock_pie.dataLabels.showBubbleSize = False

            if stock_pie.plot_area is None:
                stock_pie.plot_area = PlotArea()

            if stock_pie.plot_area.graphicalProperties is None:
                stock_pie.plot_area.graphicalProperties = GraphicalProperties()

            if stock_pie.graphical_properties is None:
                stock_pie.graphical_properties = GraphicalProperties()

            # Настраиваем фон диаграммы
            stock_pie.plot_area.graphicalProperties.noFill = True
            stock_pie.graphical_properties.gradFill = GradientFillProperties(
                gsLst=(
                    GradientStop(pos=0, srgbClr="EEF1F6"),      
                    GradientStop(pos=100000, srgbClr="C4CAE9"), 
                )
            )

            # Сдвигаем диаграмму вправо
            stock_pie.layout = Layout(
                manualLayout=ManualLayout(
                    x=1,  # Сдвиг вправо
                )
            )
            ws.add_chart(stock_pie, "D17")

            # Создаём 5 график
        if rec_data_end > rec_start_row:
            bar_chart = BarChart3D()
            bar_chart.title = "Рекомендации по распределению товара"
            bar_chart.x_axis.title = "Склад"
            bar_chart.y_axis.title = "Количество"
            bar_chart.y_axis.title.overlay = False
            bar_chart.x_axis.title.overlay = False
            bar_chart.legend.position = "r"
            bar_chart.title.overlay = False
            bar_chart.legend.position = "t"
            bar_chart.legend.overlay = False
            bar_chart.width = 30.3  # Ширина
            bar_chart.height = 15 # Высота

            bar_chart.x_axis.delete = False
            bar_chart.y_axis.delete = False

            cats = Reference(data_ws, min_col=1, min_row=rec_start_row+1, max_row=rec_data_end)  # Склад
            data = Reference(data_ws, min_col=2, max_col=5, min_row=rec_start_row, max_row=rec_data_end)  # Данные

            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)

            bar_chart.dataLabels = DataLabelList()
            bar_chart.dataLabels.showVal = True
            bar_chart.dataLabels.showCatName = False
            bar_chart.dataLabels.showSerName = False
            bar_chart.dataLabels.showLegendKey = False
            bar_chart.dataLabels.showBubbleSize = False

            # Отключаем инверсию для отрицательных значений
            for series in bar_chart.series:
                series.invertIfNegative = False

            if bar_chart.plot_area is None:
                bar_chart.plot_area = PlotArea()

            if bar_chart.plot_area.graphicalProperties is None:
                bar_chart.plot_area.graphicalProperties = GraphicalProperties()

            if bar_chart.graphical_properties is None:
                bar_chart.graphical_properties = GraphicalProperties()

            # Настраиваем фон диаграммы
            bar_chart.plot_area.graphicalProperties.noFill = True
            bar_chart.graphical_properties.gradFill = GradientFillProperties(
                gsLst=(
                    GradientStop(pos=0, srgbClr="EEF1F6"),      
                    GradientStop(pos=100000, srgbClr="C4CAE9"), 
                )
            )

            # Добавляем график на основной лист
            ws.add_chart(bar_chart, "A31")            
            
        data_ws.sheet_state = "hidden"  # скрываем лист с данными
            
            

def register_report_handler(dp: Dispatcher):
    """
    Регистрируем хендлер /my_products, который генерирует
    двухлистовый Excel-файл (первый лист группировка по категориям,
    второй лист топ-список товаров).
    """
    dp.message.register(cmd_my_products, Command("my_products"))
