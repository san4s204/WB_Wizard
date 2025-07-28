import datetime
import io
from aiogram import types, Dispatcher
from aiogram.filters import Command
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from db.database import SessionLocal
from db.models import Order, Sale, Stock, Product
from states.user_state import user_states

async def generate_excel_report_for_date(token_id: int, day_str: str) -> bytes:
    """
    Формирует Excel-отчёт (в виде байтов) за указанный день (формат YYYY-MM-DD),
    по конкретному токену (т.е. для конкретного пользователя).
    Включает те же листы, что и generate_daily_excel_report:
      1) Заказы
      2) Выкупы
      3) Отказы (is_cancel = True)
      4) Товары с нулевыми остатками
    """

    # 1) Парсим входную дату
    try:
        day_dt = datetime.datetime.strptime(day_str, "%Y-%m-%d")
    except ValueError:
        # Если формат не подошёл, выбрасываем исключение (или вернём None)
        raise ValueError(f"Некорректная дата, ожидаем YYYY-MM-DD, получили: {day_str}")

    # Определяем период: от day_dt (00:00) до day_dt + 1 day (00:00)
    date_from = day_dt
    date_to = day_dt + datetime.timedelta(days=1)

    session = SessionLocal()

    wb = Workbook()
    ws_orders: Worksheet = wb.active
    ws_orders.title = "Orders"

    # 1) Лист с заказами
    orders = (
        session.query(Order)
        .filter(
            Order.token_id == token_id,
            Order.date >= date_from,
            Order.date < date_to    # < потому что мы берём именно этот день
        )
        .all()
    )

    ws_orders.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион"])
    for cell in ws_orders[1]:
        if cell.column > 1:
            cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")

    row_index = 2
    for o in orders:
        date_val = o.date.strftime("%Y-%m-%d %H:%M:%S") if o.date else ""
        nm_id = o.nm_id or ""
        product_name = o.subject or ""
        price_val = o.price_with_disc or 0
        warehouse_val = o.warehouse_name or ""
        region_val = o.region_name or ""

        ws_orders.cell(row=row_index, column=2, value=date_val)
        ws_orders.cell(row=row_index, column=3, value=nm_id)
        ws_orders.cell(row=row_index, column=4, value=product_name)
        ws_orders.cell(row=row_index, column=5, value=price_val)
        ws_orders.cell(row=row_index, column=6, value=warehouse_val)
        ws_orders.cell(row=row_index, column=7, value=region_val)

        # Картинка
        product = session.query(Product).filter_by(nm_id=o.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                excel_img = ExcelImage(new_img_bytes)
                cell_position = f"A{row_index}"
                ws_orders.add_image(excel_img, cell_position)
                ws_orders.row_dimensions[row_index].height = 60
            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={o.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ws_orders.column_dimensions[col_letter].width = 15

    # 2) Лист с выкупами
    ws_sales = wb.create_sheet(title="Выкупы")
    sales = (
        session.query(Sale)
        .filter(
            Sale.token_id == token_id,
            Sale.date >= date_from,
            Sale.date < date_to
        )
        .all()
    )
    ws_sales.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион"])
    for cell in ws_sales[1]:
        if cell.column > 1:
            cell.fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")

    row_index = 2
    for s in sales:
        date_val = s.date.strftime("%Y-%m-%d %H:%M:%S") if s.date else ""
        nm_id = s.nm_id or ""
        product_name = s.subject or ""
        price_val = s.price_with_disc or 0
        warehouse_val = s.warehouse_name or ""
        region_val = s.region_name or ""

        ws_sales.cell(row=row_index, column=2, value=date_val)
        ws_sales.cell(row=row_index, column=3, value=nm_id)
        ws_sales.cell(row=row_index, column=4, value=product_name)
        ws_sales.cell(row=row_index, column=5, value=price_val)
        ws_sales.cell(row=row_index, column=6, value=warehouse_val)
        ws_sales.cell(row=row_index, column=7, value=region_val)

        # Картинка
        product = session.query(Product).filter_by(nm_id=s.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                excel_img = ExcelImage(new_img_bytes)
                cell_position = f"A{row_index}"
                ws_sales.add_image(excel_img, cell_position)
                ws_sales.row_dimensions[row_index].height = 60
            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={s.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ws_sales.column_dimensions[col_letter].width = 15

    # 3) Лист с отказами
    ws_cancels = wb.create_sheet(title="Отказы")
    cancels = (
        session.query(Order)
        .filter(
            Order.token_id == token_id,
            Order.is_cancel == True,
            Order.date >= date_from,
            Order.date < date_to
        )
        .all()
    )
    ws_cancels.append(["", "Дата", "Артикул", "Товар", "Цена", "Склад", "Регион", "Отменён?"])
    for cell in ws_cancels[1]:
        if cell.column > 1:
            cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    row_index = 2
    for c in cancels:
        date_val = c.date.strftime("%Y-%m-%d %H:%M:%S") if c.date else ""
        nm_id = c.nm_id or ""
        product_name = c.subject or ""
        price_val = c.price_with_disc or 0
        warehouse_val = c.warehouse_name or ""
        region_val = c.region_name or ""
        is_cancel_val = "Да" if c.is_cancel else "Нет"

        ws_cancels.cell(row=row_index, column=2, value=date_val)
        ws_cancels.cell(row=row_index, column=3, value=nm_id)
        ws_cancels.cell(row=row_index, column=4, value=product_name)
        ws_cancels.cell(row=row_index, column=5, value=price_val)
        ws_cancels.cell(row=row_index, column=6, value=warehouse_val)
        ws_cancels.cell(row=row_index, column=7, value=region_val)
        ws_cancels.cell(row=row_index, column=8, value=is_cancel_val)

        product = session.query(Product).filter_by(nm_id=c.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                excel_img = ExcelImage(new_img_bytes)
                cell_position = f"A{row_index}"
                ws_cancels.add_image(excel_img, cell_position)
                ws_cancels.row_dimensions[row_index].height = 60
            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={c.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 9):
        col_letter = get_column_letter(col)
        ws_cancels.column_dimensions[col_letter].width = 15

    # 4) Лист с нулевыми остатками
    ws_out_of_stock = wb.create_sheet(title="Отсутствие товаров")
    zero_stocks = (
        session.query(Stock)
        .filter(
            Stock.token_id == token_id,
            Stock.quantity == 0
        )
        .all()
    )
    ws_out_of_stock.append(["", "Артикул", "Склад", "Остаток", "Дата обновления"])
    for cell in ws_out_of_stock[1]:
        cell.fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")

    row_index = 2
    for z in zero_stocks:
        nm_id = z.nm_id or ""
        warehouse_val = z.warehouseName or ""
        quantity_val = z.quantity or 0
        date_val = z.last_change_date.strftime("%Y-%m-%d %H:%M:%S") if z.last_change_date else ""

        ws_out_of_stock.cell(row=row_index, column=2, value=nm_id)
        ws_out_of_stock.cell(row=row_index, column=3, value=warehouse_val)
        ws_out_of_stock.cell(row=row_index, column=4, value=quantity_val)
        ws_out_of_stock.cell(row=row_index, column=5, value=date_val)

        product = session.query(Product).filter_by(nm_id=z.nm_id, token_id=token_id).first()
        if product and product.resize_img:
            try:
                img_data = io.BytesIO(product.resize_img)
                pil_img = PILImage.open(img_data)
                pil_img = pil_img.resize((80, 80), PILImage.Resampling.LANCZOS)
                new_img_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")
                pil_img.save(new_img_bytes, format="JPEG", optimize=True, quality=70)
                new_img_bytes.seek(0)

                excel_img = ExcelImage(new_img_bytes)
                cell_position = f"A{row_index}"
                ws_out_of_stock.add_image(excel_img, cell_position)
                ws_out_of_stock.row_dimensions[row_index].height = 60
            except Exception as exc:
                print(f"Не удалось вставить картинку nm_id={z.nm_id}: {exc}")

        row_index += 1

    for col in range(1, 6):
        col_letter = get_column_letter(col)
        ws_out_of_stock.column_dimensions[col_letter].width = 15

    # 5) Стили (границы и выравнивание)
    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def apply_styles_to_worksheet(ws):
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment

    for sheet in [ws_orders, ws_sales, ws_cancels, ws_out_of_stock]:
        apply_styles_to_worksheet(sheet)

    session.close()

    # Преобразуем Workbook в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def cmd_report_for_day(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id] = {"state": "await_report_date"}
    await message.answer(
        "Введите дату для формирования отчёта в формате <b>YYYY-MM-DD</b>.\n"
        "Например: <code>2025-03-25</code>\n\n"
        "❌ Чтобы отменить ввод — введите /cancel",
        parse_mode="HTML"
    )
def register_repots_for_day_handler(dp: Dispatcher):
    dp.message.register(cmd_report_for_day, Command("report_for_day"))