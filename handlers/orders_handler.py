import io
import datetime
from aiogram import types, Dispatcher
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from core.sub import user_has_role
from db.database import SessionLocal
from db.models import User, Order, Product
from PIL import Image as PILImage
from collections import defaultdict

MAX_TELEGRAM_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

async def cmd_orders(message: types.Message, user_id:int = None, days:int = 0):
    """
    Хендлер: /orders_report <days>
    Пример: /orders_report 30
    Если дней не указано, берём 7 по умолчанию.
    Генерируем Excel, где:
      - Колонки: A(Артикул + картинка), B(Сумма), C(Количество), D(Сумма отказов), E(Кол-во отказов), F(Размер).
      - Строки: по каждому nm_id и каждому techSize.
    """

    # 1) Период из аргумента
    args = message.text.split()
    if len(args) > 1 and args[1].isdigit():
        days = int(args[1])
    else:
        days = days

    if days == 0:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="7 дней", callback_data="orders 7")],
                [InlineKeyboardButton(text="30 дней", callback_data="orders 30")],
                [InlineKeyboardButton(text="90 дней", callback_data="orders 90")]
            ]
        )
        await message.answer("За какой период вывести заказы?", reply_markup=kb)
        return  # Прерываем, отчёт пока не генерируем

    if user_id is None:
        user_id = message.from_user.id

    # 2) Ищем user + token_id
    session = SessionLocal()
    db_user = session.query(User).filter_by(telegram_id=str(user_id)).first()
    if not db_user or not db_user.token_id:
        print(db_user)
        await message.answer("Нет привязанного токена. Сначала /start и пришлите токен.")
        session.close()
        return


    if days == 7:
        allowed_roles = ["base", "advanced", "test", "super"]
    elif days in (30, 90):
        allowed_roles = ["advanced", "test", "super"]
    else:
        # Если пользователь ввёл другие числа, то, например, только super
        allowed_roles = ["super"]

    # 3) Проверяем роль через user_has_role
    has_access = user_has_role(session, str(user_id), allowed_roles)
    if not has_access:
        await message.answer(
            f"У вас нет доступа к просмотру заказов за {days} дней.\n"
            f"Доступны только роли: {', '.join(allowed_roles)}."
        )
        session.close()
        return

    # 3) Получаем заказы за нужный период
    date_from = datetime.datetime.utcnow() - datetime.timedelta(days=days)
    orders = (
        session.query(Order)
        .filter(Order.token_id == db_user.token_id)
        .filter(Order.date >= date_from)
        .all()
    )
    session.close()

    if not orders:
        await message.answer(f"За {days} дней заказов нет.")
        return

    # 4) Создаём Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    headers = ["Артикул", "Кол-во заказов", "Сумма заказов", "Кол-во отказов", "Сумма отказов", "Размер"]

    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = Font(bold=True, color="FF000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ширина столбцов
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12

    # Цвета для колонок (кроме A)
    col_fills = {
        2: PatternFill(start_color="FFC4D79B", end_color="FFC4D79B", fill_type="solid"),  # B
        3: PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),  # C
        4: PatternFill(start_color="FFE26B0A", end_color="FFE26B0A", fill_type="solid"),  # D
        5: PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),  # E
        6: PatternFill(start_color="FF60497A", end_color="FF60497A", fill_type="solid"),  # F
    }

    # Тонкие границы
    thin_border = Border(
        left=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
        top=Side(border_style="thin", color="FF000000"),
        bottom=Side(border_style="thin", color="FF000000"),
    )

    # Красим шапку, ставим границы
    for col_idx in range(1, 7):
        hdr_cell = ws.cell(row=1, column=col_idx)
        if col_idx in col_fills:
            hdr_cell.fill = col_fills[col_idx]
        hdr_cell.border = thin_border

    # 5) Группируем заказы (nm_id, techSize)
    data_map = defaultdict(list)
    nm_ids = set()
    for o in orders:
        key = (o.nm_id, o.techSize or "")
        data_map[key].append(o)
        nm_ids.add(o.nm_id)

    # Подтягиваем продукты
    session2 = SessionLocal()
    products_db = session2.query(Product).filter(Product.nm_id.in_(nm_ids)).all()
    session2.close()
    products_map = {p.nm_id: p for p in products_db}

    inserted_images_for = set()  # чтобы не вставлять картинку повторно
    sorted_keys = sorted(data_map.keys(), key=lambda x: (x[0], x[1]))

    current_row = 2
    last_nm_id = None

    for (nm_id, size_val) in sorted_keys:
        # Если начался новый nm_id (артикул) – сделаем отступ (пропустим строку).
        if last_nm_id is not None and nm_id != last_nm_id:
            current_row += 5

        orders_for_key = data_map[(nm_id, size_val)]
        count_orders = sum(1 for o in orders_for_key if not o.is_cancel)
        sum_orders = sum((o.price_with_disc or 0.0) for o in orders_for_key if not o.is_cancel)
        count_cancel = sum(1 for o in orders_for_key if o.is_cancel)
        sum_cancel = sum((o.price_with_disc or 0.0) for o in orders_for_key if o.is_cancel)

        # (1) Пишем артикул в текущую строку (колонку A)
        cell_a = ws.cell(row=current_row, column=1, value=str(nm_id))
        cell_b = ws.cell(row=current_row, column=2, value=int(count_orders))
        cell_c = ws.cell(row=current_row, column=3, value=float(sum_orders))
        cell_d = ws.cell(row=current_row, column=4, value=int(count_cancel))
        cell_e = ws.cell(row=current_row, column=5, value=float(sum_cancel))
        cell_f = ws.cell(row=current_row, column=6, value=size_val)

        # Стили
        for col_idx in range(1, 7):
            cell_obj = ws.cell(row=current_row, column=col_idx)
            if col_idx in col_fills:
                cell_obj.fill = col_fills[col_idx]
            cell_obj.border = thin_border
            if col_idx == 3 or col_idx == 5:
                cell_obj.number_format = '#,##0.00'

        # (2) Если для этого nm_id ещё не вставляли картинку – вставляем
        if nm_id not in inserted_images_for:
            product = products_map.get(nm_id)
            if product and product.resize_img:
                try:
                    img_bytes = io.BytesIO(product.resize_img)
                    pil_img = PILImage.open(img_bytes)
                    pil_img = pil_img.resize((168, 140), PILImage.Resampling.LANCZOS)
                    out_bytes = io.BytesIO()
                    pil_img = pil_img.convert("RGB")
                    pil_img.save(out_bytes, format="JPEG", optimize=True, quality=70)
                    out_bytes.seek(0)
                    excel_img = ExcelImage(out_bytes)

                    # Ставим картинку в ту же колонку A, но на строку ниже
                    img_row = current_row + 1
                    anchor_cell = f"A{img_row}"
                    ws.add_image(excel_img, anchor_cell)

                    # Чтобы не обрезалось, поднимем высоту строки (img_row)
                    # 130 px ~ ~100 пунктов, но можно и 80..90, кто как любит


                    # Можно отступить ещё строки, чтобы следующий товар не наехал
                    # Допустим, пропустим ещё 2 строки. Итого товар занимал 2 строки
                    # + 2 пропущенные = 4 строки на 1 nm_id

                except Exception as e:
                    print("Ошибка при вставке картинки:", e)

            inserted_images_for.add(nm_id)

        # (3) Переходим на следующую строку для следующей записи
        current_row += 1
        last_nm_id = nm_id

    # Теперь преобразуем workbook в байты, и отсылаем пользователю
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    workbook_bytes = output.getvalue()

    file_size = len(workbook_bytes)
    if file_size > MAX_TELEGRAM_FILE_SIZE:
        await message.answer("Извините, итоговый файл слишком большой для отправки.")
        return

    doc = types.BufferedInputFile(workbook_bytes, filename=f"Заказы за {days} дней.xlsx")
    await message.answer_document(document=doc, caption=f"Отчёт по заказам за {days} дней.")

def register_orders_handler(dp: Dispatcher):
    dp.message.register(cmd_orders, Command("orders"))
