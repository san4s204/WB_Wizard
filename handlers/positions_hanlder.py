import io
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, PatternFill, Font
from aiogram.filters import Command
from openpyxl.utils import get_column_letter
from aiogram.types import BufferedInputFile
from PIL import Image as PILImage
from db.models import DestCity, ProductPositions, Product, User
from aiogram import types, Dispatcher
from sqlalchemy import func
from db.database import SessionLocal
from collections import defaultdict

async def cmd_positions(message: types.Message, user_id: int = None):

    session = SessionLocal()

    if user_id is None:
        user_id = message.from_user.id

    db_user = session.query(User).filter_by(telegram_id=str(user_id)).first()
    if not db_user or not db_user.token_id:
        await message.answer("Нет привязанного токена. Сначала /start и пришлите токен.")
        session.close()
        return

    

    # Создаём Workbook
    wb = Workbook()

    # Удаляем дефолтный лист "Sheet"
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Генерируем отчёт
    generate_positions_report(session, wb)
    
    # Генерирем листы с динамикой позиций по городам
    generate_dynamic_positions_report(session, wb)

    session.close()

    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    workbook_bytes = output.getvalue()

    file_size = len(workbook_bytes)
    if file_size > 50 * 1024 * 1024:  # 50 MB
        await message.answer("Слишком большой Excel для отправки!")
        return

    doc = BufferedInputFile(workbook_bytes, filename="positions_report.xlsx")
    await message.answer_document(document=doc, caption="Отчёт позиций")

def get_default_period(session) -> tuple:
    """
    Если период не задан, вычисляет его:
      - start_date: минимальное значение check_dt среди записей в ProductPositions.
      - end_date: start_date + 3 месяца (приблизительно 90 дней).
    Возвращает кортеж (start_date, end_date) типа (datetime.date, datetime.date).
    """
    min_dt = session.query(func.min(ProductPositions.check_dt)).scalar()
    if min_dt is None:
        start_date = datetime.date.today()
    else:
        start_date = min_dt.date()
    end_date = start_date + datetime.timedelta(days=90)
    return start_date, end_date

def generate_positions_report(session, wb):
    """
    Создаёт лист "Positions" в книге wb:
      - Собирает данные из product_positions (page, position, request_count, query_text),
        связав с таблицами Product (чтобы достать nm_id, resize_img) и DestCity (чтобы
        получить названия городов).
      - Выводит колонки:
         A: Товар (название + картинка)
         B: Частотность (месяц)
         C: Ключевое слово
         D..: Города (page–pos), напр. "25–8"
    """

    ws = wb.create_sheet(title="Positions")

    # 1) Список всех городов (для упрощения — сортируем по city)
    all_cities = session.query(DestCity).order_by(DestCity.city).all()
    city_count = len(all_cities)

    # 2) Пишем заголовок в строчку:
    # A1 = "Товар", B1 = "Частотность (месяц)", C1 = "Ключевое слово"
    ws["A1"] = "Товар"
    ws["B1"] = "Частотность запросов(месяц)"
    ws["C1"] = "Ключевое слово"

    # Начиная с D, E, F... — города
    col_start = 4
    for i, city in enumerate(all_cities):
        col_idx = col_start + i
        ws.cell(row=1, column=col_idx, value=city.city)

    # Стили шапки
    header_fill = PatternFill(start_color="FFB0E0E6", end_color="FFB0E0E6", fill_type="solid")
    header_font = Font(bold=True, color="FF000000")

    max_col = col_start + city_count - 1
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Немного поднастроим ширину
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 35
    # города тоже можно поднастроить
    for i in range(city_count):
        letter = get_column_letter(col_start + i)
        ws.column_dimensions[letter].width = 15

    # 3) Собираем товары, у которых есть записи в product_positions
    #    (или просто все товары, фильтруем потом).

    # Список distinct nm_id
    nm_ids_in_positions = session.query(ProductPositions.nm_id).distinct().all()
    nm_ids_in_positions = [nm_id for (nm_id,) in nm_ids_in_positions]

    products = (
        session.query(Product)
        .filter(Product.nm_id.in_(nm_ids_in_positions))
        .all()
    )

    current_row = 2

    for product in products:
        nm_id = product.nm_id

        # Вытаскиваем картинку (resize_img), если есть
        resize_img_bytes = product.resize_img  # LargeBinary
        # Создадим картинку excel_img, если не None
        excel_img = None
        if resize_img_bytes:
            try:
                pil_img = PILImage.open(io.BytesIO(resize_img_bytes))
                # Подогнать под (150x150)?
                pil_img = pil_img.resize((182, 160), PILImage.Resampling.LANCZOS)
                out_bytes = io.BytesIO()
                pil_img = pil_img.convert("RGB")  # для JPEG нужно RGB
                pil_img.save(out_bytes, format="JPEG", optimize=True, quality=70)
                out_bytes.seek(0)
                excel_img = ExcelImage(out_bytes)
            except Exception as e:
                print(f"Не удалось загрузить/преобразовать картинку nm_id={nm_id}: {e}")

        # 4) Собираем все записи product_positions для этого nm_id,
        #    группируем по query_text (одна строка => 1 query_text => "Частотность" => города).
        positions_data = session.query(ProductPositions).filter_by(nm_id=nm_id).all()
        if not positions_data:
            continue

        # Соберём словарь: key= (query_text, request_count) => city -> (page, pos)
        from collections import defaultdict
        # можно хранить query_info[(query_text, freq)][city_id] = (page, pos)
        query_info = defaultdict(lambda: {})

        for pp in positions_data:
            key = (pp.query_text, pp.request_count)
            query_info[key][pp.city_id] = (pp.page, pp.position)

         # Если меньше 5 ключевых слов => пропускаем товар
        if len(query_info) < 5:
            continue

        # Перед тем как писать строки, сделаем "шапку" для товара:
        # - вставим картинку в A{current_row}
        # - объединим A..C чтобы вписать nm_id или supplier_article
        #   (или subject_name, brand_name — на ваш вкус)
        # Пример: "WB страница: https://www.wildberries.ru/catalog/<nm_id>/detail.aspx"
        # + subject_name
        link_url = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"
        product_title = f'=HYPERLINK("{link_url}","WB страница nm_id: {nm_id}")'
        
        # Объединим A..C (3 колонки)
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=3
        )
        cell_title = ws.cell(row=current_row, column=1)
        cell_title.value = product_title
        cell.hyperlink = link_url
        cell.style = "Hyperlink"


        cell_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_title.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell_title.font = Font(bold=True)

        # Вставляем картинку, если есть
        if excel_img:
            ws.add_image(excel_img, f"A{current_row+1}")

        current_row += 1  # отступаем одну строку

        # Пишем строки => по каждому (query_text, freq), одна строка
        for (qtext, freq) in query_info.keys():
            if freq == 0:
                freq_value = "<100"
            else:
                freq_value = freq

            # Частотность в B
            cell_freq = ws.cell(row=current_row, column=2, value=freq_value)
            cell_freq.alignment = Alignment(horizontal="right", vertical="center")
            # Ключевое слово в C
            ws.cell(row=current_row, column=3, value=qtext)

            # Далее города: D.. col_start + i => all_cities[i].id
            for i, city in enumerate(all_cities):
                city_id = city.id
                page_pos = query_info[(qtext, freq)].get(city_id, (None, None))
                page, pos = page_pos
                col_idx = col_start + i

                cell_val = ">30"
                cell_fill = PatternFill(start_color="FF9F9FFF", end_color="FF9F9FFF", fill_type="solid")

                if page is not None and pos is not None:
                    cell_val = f"{page}–{pos}"
                    if page <= 10:
                        # зелёный: “light green” ARGB=FF90EE90
                        cell_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
                    else:
                        # красный: “light red” ARGB=FFFFA07A
                        cell_fill = PatternFill(start_color="FFFFA07A", end_color="FFFFA07A", fill_type="solid")

                cell_obj = ws.cell(row=current_row, column=col_idx, value=cell_val)
                # Если есть fill
                if cell_fill:
                    cell_obj.fill = cell_fill
            current_row += 1

        # Небольшой отступ между товарами
        current_row += 1

    print("[generate_positions_report] Завершено.")

def generate_dynamic_positions_report(session, wb, start_date: datetime.date = None, end_date: datetime.date = None):
    """
    Для каждого города (из таблицы DestCity) создаёт отдельный лист.
    В листе выводится динамика позиций товаров по дням за период [start_date, end_date].
    
    Структура листа:
      - Первая строка: заголовки: "Товар" (с изображением и гиперссылкой), затем по дням (формат "ДД.MM").
      - Каждая строка: для одного товара. В первой колонке – данные о товаре, затем для каждого дня:
          Если для данного товара (и города) есть запись, вычисляем global_rank = (page-1)*100 + pos.
          Если есть предыдущий день, вычисляем разницу:
            * Если улучшение (global_rank уменьшился) – добавляем "↑<diff>" и заливаем зелёным.
            * Если ухудшение – добавляем "↓<diff>" и заливаем красным.
          Если записи нет – выводим "-".
    """

     # Если период не задан, вычисляем его
    if start_date is None or end_date is None:
        start_date, end_date = get_default_period(session)
        print(f"[INFO] Используем период по умолчанию: {start_date} - {end_date}")


    # Получаем список дат в нужном диапазоне (каждый день)
    days_list = []
    current_day = start_date
    while current_day <= end_date:
        days_list.append(current_day)
        current_day += datetime.timedelta(days=1)
    

    # Получаем все города
    cities = session.query(DestCity).order_by(DestCity.city).all()
    
    for city in cities:
        # Собираем все ProductPositions для данного города в диапазоне дат
        pps_city = session.query(ProductPositions).filter(
            ProductPositions.city_id == city.id,
            ProductPositions.check_dt >= datetime.datetime.combine(start_date, datetime.time.min),
            ProductPositions.check_dt <= datetime.datetime.combine(end_date, datetime.time.max)
        ).all()
        
        # Если нет данных для города — просто создаём лист и пропускаем
        if not pps_city:
            ws = wb.create_sheet(title=city.city[:31])
            continue

        # Собираем уникальные даты, по которым есть данные
        days_with_data = set()
        for pp in pps_city:
            days_with_data.add(pp.check_dt.date())
        # Сортируем даты, чтобы идти по порядку
        days_list = sorted(days_with_data)

        # Создаём лист для города
        ws = wb.create_sheet(title=city.city[:31])

        # Шапка (A1, B1, C1)
        fill_head = PatternFill(start_color="FFB0E0E6", end_color="FFB0E0E6", fill_type="solid")
        ws["A1"] = "Товар"
        ws["A1"].fill = fill_head
        ws["B1"] = "Частотность запросов(месяц)"
        ws["B1"].fill = fill_head
        ws["C1"] = "Ключевое слово"
        ws["C1"].fill = fill_head

        # Нумеруем столбцы, начиная с 4-го (D)
        col_start = 4
        for i, d in enumerate(days_list):
            col_idx = col_start + i
            cell = ws.cell(row=1, column=col_idx)
            cell.value = d.strftime("%d.%m.%Y")
            fill_date = PatternFill(start_color="FFBF43EB", end_color="FFBF43EB", fill_type="solid")
            cell.fill = fill_date
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)

        # Устанавливаем ширину столбцов
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        for i in range(len(days_list)):
            col_letter = get_column_letter(col_start + i)
            ws.column_dimensions[col_letter].width = 15

        # nm_ids (товары, у которых в этом городе есть позиции)
        nm_ids = set(pp.nm_id for pp in pps_city)
        products = session.query(Product).filter(Product.nm_id.in_(nm_ids)).all()

        current_row = 2

        for product in products:
            nm_id = product.nm_id

            # Собираем позиции именно этого товара за нужный период и в этом городе
            pps_for_product = [pp for pp in pps_city if pp.nm_id == nm_id]
            if not pps_for_product:
                continue

            # Группируем pps по (query_text, request_count)
            query_map = defaultdict(list)
            for pp in pps_for_product:
                key = (pp.query_text, pp.request_count)
                query_map[key].append(pp)

            # Если у товара <5 ключевых слов => пропускаем (логика по вашему условию)
            if len(query_map) < 5:
                continue

            # Вставляем "шапку" товара: объединим A..C
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            cell_title = ws.cell(row=current_row, column=1)

            link_url = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"
            cell_title.value = f'=HYPERLINK("{link_url}","WB страница nm_id: {nm_id}")'
            cell_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_title.font = Font(bold=True, color="FF000000")
            cell_title.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

            # Пробуем вставить картинку (если есть)
            excel_img = None
            if product.resize_img:
                try:
                    pil_img = PILImage.open(io.BytesIO(product.resize_img))
                    pil_img = pil_img.resize((182, 160), PILImage.Resampling.LANCZOS)
                    out_bytes = io.BytesIO()
                    pil_img = pil_img.convert("RGB")
                    pil_img.save(out_bytes, format="JPEG", optimize=True, quality=70)
                    out_bytes.seek(0)
                    excel_img = ExcelImage(out_bytes)
                except Exception as e:
                    print(f"[generate_dynamic_positions_report] Ошибка с картинкой nm_id={nm_id}: {e}")

            if excel_img:
                ws.add_image(excel_img, f"A{current_row+1}")

            current_row += 1  # Отступ на 1 строку

            # Для каждого (qtext, freq) одна строка с позициями
            for (qtext, freq), pp_list in query_map.items():
                if freq == 0:
                    freq_value = "<100"
                else:
                    freq_value = freq

                cell_freq = ws.cell(row=current_row, column=2, value=freq_value)
                cell_freq.alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=current_row, column=3, value=qtext)

                # Готовим словарь date->(pp)
                # Будем брать самую позднюю запись для каждого дня
                daily_positions = {}
                for pp in pp_list:
                    d = pp.check_dt.date()
                    if d not in daily_positions or pp.check_dt > daily_positions[d].check_dt:
                        daily_positions[d] = pp

                prev_rank = None
                # Идём по days_list (только те дни, где есть хоть какие-то записи в городе)
                for i_day, day_obj in enumerate(days_list):
                    col_idx = col_start + i_day
                    cell_obj = ws.cell(row=current_row, column=col_idx)
                    # Если в этот день есть запись для данного (qtext,freq)
                    if day_obj in daily_positions:
                        pp = daily_positions[day_obj]
                        if pp.page is not None and pp.position is not None:
                            rank = (pp.page - 1) * 100 + pp.position
                            diff_text = ""
                            fill = None
                            if prev_rank is not None:
                                diff = prev_rank - rank
                                if diff > 0:
                                    # Улучшение
                                    diff_text = f" ↑{diff}"
                                    fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
                                elif diff < 0:
                                    # Ухудшение
                                    diff_text = f" ↓{abs(diff)}"
                                    fill = PatternFill(start_color="FFFFA07A", end_color="FFFFA07A", fill_type="solid")
                                else:
                                    # Нулевая динамика
                                    diff_text = " →0"
                                    fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
                            else:
                                diff_text = ""
                                fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")  # зелёный

                            cell_obj.value = f"{pp.page}-{pp.position}{diff_text}"
                            if fill:
                                cell_obj.fill = fill
                            prev_rank = rank
                        else:
                            cell_obj.value = ">30"
                            fill = PatternFill(start_color="FF9F9FFF", end_color="FF9F9FFF", fill_type="solid")
                            cell_obj.fill = fill
                            # prev_rank не сбрасываем
                    else:
                        cell_obj.value = ">30"
                        fill = PatternFill(start_color="FF9F9FFF", end_color="FF9F9FFF", fill_type="solid")
                        cell_obj.fill = fill
                        # prev_rank не сбрасываем

                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

            # Отступ между товарами
            current_row += 1

    print("[generate_dynamic_positions_report] Завершено.")

def register_positions_handlers(dp: Dispatcher):
    """
    Регистрация хендлеров для работы с отчётом позиций.
    """
    dp.message.register(cmd_positions, Command("positions"))